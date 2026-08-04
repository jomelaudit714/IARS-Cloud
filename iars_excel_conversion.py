from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import base64
from pathlib import Path
import hashlib
import re
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


WAREHOUSE_OUTPUT_HEADERS = [
    "id",
    "login_date",
    "sap_no",
    "sap_date",
    "locations",
    "stock_status",
    "task",
    "product",
    "uom",
    "record_qty",
    "count_qty",
    "remarks",
    "auditor_name",
]

WAREHOUSE_TEMPLATE_PATH = (
    Path(__file__).resolve().parent / "assets" / "warehouse_conversion_template.xlsx"
)
# Embedded approved template fallback. This keeps the module operational even when
# deployment tools copy Python files but omit the assets folder.
WAREHOUSE_TEMPLATE_XLSX_BASE64 = (
    'UEsDBBQAAAAIACpK/1xGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYp'
    'bYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3'
    'sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIACpK/1ypO90dEQEAAGYCAAARAAAA'
    'ZG9jUHJvcHMvY29yZS54bWzNkkFqwzAQRa8StLcl2cUhwvGiLV2UBgoJtHQn5IkjallCmuLk9pVdx2lID9Dl/Pl68wdNqZxQ1sOr'
    'tw48agiLo2m7IJRbkwOiE5QGdQAjQxodXWzurTcSY+kb6qT6lA3QjLGCGkBZS5R0ACZuJpKqrJVQHiRaP+FrNePdl29HWK0otGCg'
    'w0B5yimpnq2BdrGVHWrZ2JJeMAMSwZvwI0A9c0f1T/jYoWRyHoOeXX3fp30++uImnL5vXrbj0onuAspOQXwVtMCTgzU5T37LHx53'
    'T6TKWFYkjCd8tWO5YHciKz6GrFf5LoGNrfVe/4PEyyTnO7YSfCky9ivxOWBVxuNoZcDNJNyfbn7k1jFq1ydVfQNQSwMEFAAAAAgA'
    'Kkr/XMEXEL6SBgAAxiAAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7VnNb9s2FL8P2P8g6O5KtiV/BHUKW7b7lTRB43bokZZpizEl'
    'GiSVxCgKDO1plwEDumGXAbvtMAwrsAIrdtkfE6DF1v0Re5K/RJtqkzYtOiwOYJPU7z3++N7j44t49dpJSI0jzAVhUcMsXrFNA0c+'
    'G5Bo1DDv9bqFmmkIiaIBoizCDXOKhXlt+/PPrqItGeAQGyAfiS3UMAMpJ1uWJXwYRuIKm+AIng0ZD5GELh9ZA46OQW9IrZJtV6wQ'
    'kcg0IhSC2r3hkPjY6CUqze2F8g6Fr0iKZMCn/MBPZ8xKpNjBuJj8iKnwKDeOEG2YMM+AHffwiTQNioSEBw3TTj+mtX3VWgpRmSOb'
    'keumn7ncXGAwLqVyfNRfCjqO61SaS/2lmf5NXKfaqXQqS30pAPk+rLSo0Vktec4cmwHNmhrd7Wq7XFTwGf3lDXzTTf4UfHmFdzbw'
    '3a63smEGNGu6G3i3VW+1Vf3uCl/ZwFftZtupKvgUFFASjTfQtlspe4vVLiFDRm9o4XXX6VZLc/gKZWWiayYfybxYC9Eh410ApM5F'
    'kkSGnE7wEPmA8xAlfU6MHTIKIPAmKGIChu2S3bXL8J38OWkr9SjawigjPRvyxcZQwscQPicT2TBvgVYzA3n54sXp4+enj38/ffLk'
    '9PGv87k35W6gaJSVe/3TN//88KXx928/vn76rR4vsvhXv3z16o8/36ReKrS+e/bq+bOX33/9189PNfAmR/0svEdCLIw7+Ni4y0JY'
    'oGYC3Ofnk+gFiCgSKACkBtiRgQK8M0VUh2th1YT3OWQKHfB6fKhwPQh4LIkGeDsIFeAuY7TFuHY5t5O5ssuJo5F+ch5ncXcROtLN'
    '7a05uBNPIOSJTqUXYIXmPgVvoxGOsDSSZ2yMsUbsASGKXXeJz5lgQ2k8IEYLEa1JeqQv9UI3SAh+meoIgqsV2+zeN1qM6tS38ZGK'
    'hG2BqE4lpooZr6NYolDLGIU0i9xBMtCRPJhyXzG4kODpEabM6AywEDqZPT5V6N6GDKN3+y6dhiqSSzLWIXcQY1lkm429AIUTLWcS'
    'BVnsTTGGEEXGPpNaEkzdIUkf/ICiXHffJ1ieb1vfgwykD5DkScx1WwIzdT9O6RBhnfImD5Xs2uREGx2teKSE9g7GFB2jAcbGvZs6'
    'PJswPelbAWSVG1hnm1tIjdWkH2EBZVJS12gcS4QSsgd4xHL47E7XEs8URSHieZrvjNWQ6cApp02le9QfK6mU8GTT6knsiRCdSet+'
    'gJSwSvpCH69THp13j4HM4TvI4HPLQGI/s216iGJ9wPQQFBi6dAsisV4k2U6pWKyVG6qbduUGa63eCUn01uJnrexxP07Z88EKnosv'
    'dfJSynqBk4f7D5Y1bRRH+xhOksuq5rKq+T9WNXl7+bKWuaxlLmuZj1bLrMoXK/uWJ9US5r7yGRJKD+SU4h2RFj4C9v6gC4NpJxVa'
    'vmGaBNCcT6fgRhylbYMz+QWRwUGAJjBNMZ1hJOaqR8KYMAGlk5mrOy294nCXDWajxeLipSYIILkah9JrMQ6FmpyNVqqrt3dL9Wlv'
    'JLIE3FTp2UlkJlNJlDUkquWzkSjaF8WirmFRK76JhZXxChxOBkreh7vOjBGEG4T0IPHTTH7h3Qv3dJ4x1WWXNMurOxfmaYVEJtxU'
    'EpkwDODwWB++YF/X63pXl7Q0qrUP4WtrMzfQSO0Zx7Dnyi6o8dGkYQ7hnyZohhPQJ5JMhegoapi+nBv6XTLLhAvZRiKYwdJHs/WH'
    'RGJuUBJCrGfdQKMVt2Kpan+65Or2p2c5a93JeDjEvswZWXXh2UyJ9ul7gpMOi4H0QTA4Nvo05ncRGMqtFhMDDoiQS2sOCM8E98qK'
    'a+lqvhWVy5bVFkV0EqD5iZJN5jN42l7SyawjZbq+Kktnwv6oexGn7tuF1pJmzgFSzc1iH+6Qz7Aq61m52lxXr9lvPiXe/0DIUKvp'
    'qZX11PLOjgssCDLTVXLsVsr15nueButRa2XqyrS3cavN+ocQ+W2oVmMqxezl2AmU397iPnKWCdLRRXY5kUbMScN8aLtNxyu5XsGu'
    'uZ2CU3bsQs1tlgtN1y0XO27RbrdKj8AoMgiL7mzuLvyzT6fzS/t0fOPiPlyU2ld8FlosrYOtVDi9uC+W8i/uDQKWeVgpdevleqtS'
    'qJeb3YLTbtUKda/SKrQrXrXdbXturd59ZBpHKdhplj2n0qkVKkXPKzgVO6FfqxeqTqnUdKrNWsdpPprbGla++F2YN+W1/S9QSwME'
    'FAAAAAgAKkr/XKQpsLjxAgAAKQoAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWydVm1zojAQ/isMP0BelBdv1JlWa9u7dqbT'
    'zt19dCJEzQiEJktt79dfEkApINR+0exunmezy0PYyYGyPd9hDNp7HCV8qu8A0h+GwYMdjhEf0BQnIrKhLEYgTLY1eMowChUojgzb'
    'NF0jRiTRZxPle2KzCc0gIgl+YhrP4hixj2sc0cNUt/TS8Uy2O5AOYzZJ0Ra/YPidPjFhGUeWkMQ44YQmGsObqX5l/Xi05X614Q/B'
    'B15Za4DWLzjCAeBQJZKVrSndy+C9cJnygGqDZETi7w3PcRRN9bktjvWqcsxVAuPIWl2X2ZaqF6K2NeJ4TqO/JITdVPd1LcQblEVQ'
    '8Q18xxm5vuccg8/0cIeL4h2ZLKARV7/aIQd5A8czh5YtMGvMYUlUn7Qg40Djglg2Ej4iPNVHuhaTRHli9F50tMJmmQPbdyzHvZDP'
    'LvjsOp8z8CxzPPQuoxsWdMPm8Ua24/mXljsq+EZ1Pvt0vB4Kp6Bw6hTe9zrmFnxuo8ReqFdAvRp0OPped/yCz6/x+Sdp9TCMC4Zx'
    'vRirIuk+SZqlJs1mi7/OclR2Q9rWwBqZ7heqsUo1Ww05Vx92H0spYquuYv+CckrlWu7QV+o18htA3S0LBGg2YfSgMQkWSeTiStIo'
    'CvH0SSKv1hdgIkoEDmYknBggeKRlBAXmuhsT0S1JViEC3IKdd2M5SlcJbcEt+nFnMt70nTZA8trmLdBlT1KgwX7FAUHWhr7tRgPi'
    '+xbUXTcqZTTMAmgB3ncDMxq3gH52gxgOKAtXr/DRgv3VjQ1olsAZ6ENfWvEd37f19LEbiLKQAGWrBMU1KRhC+Ef126X6r21F5ys6'
    'OWacdJpHnGZkcRZzk0fGzcjyLNutXSnoU+Qhj7jNyGMeGX6K5AUalVddzj6PiImXkWsR3oi95kB8DVg+Iag10FSt5CeAgrhbSmsn'
    'xjDMpCWybCiF0jjNVFmqUUZwAur1meopZcAQAQEW/n9UBKJFSuR0pL1hBiQ42vJmOk6Hs/9QSwMEFAAAAAgAKkr/XIxtW23vAgAA'
    'KBAAAA0AAAB4bC9zdHlsZXMueG1s7Vjfa9swEP5XjN47/2pNPOI8zBAYbKPQPuyhL0osJwLJ8mylJPvrp5Ncx2102bqFsY05BEv6'
    '/N13d76TQua9Pgh2t2VMB3spmr4gW63bt2HYr7dM0v6NalljkFp1kmoz7TZh33aMVj2QpAiTKMpCSXlDFvNmJ5dS98Fa7RpdkGRc'
    'CtztfVWQOLsmgTNXqooV5GCuhyspH66qioRexs0pwxCG58NBdTGvVXMUvyFuwRikkgWPVBSkpIKvOg6s9ZZ2vYnarifpDNZqKrk4'
    'DEv2ISVUF2iTCaMaw0r/1cGxm0GSBtuSN6qz/jjVM9r/dS6q021WBVkuI3u9TszeoHK4EGPlpMQtLOYt1Zp1zdJMLMcunkDBML4/'
    'tCbcTUcPcXJDfpjQK8ErkNyU03BcQGBmNQC8qdiemYYwHQTWJxZHLXszAa1UV7FuDCkmT0uLuWC1NvSOb7Zw16oFDaW1kmZQcbpR'
    'DbXxPjGGgTG7ZkLcwZbxuX7W5ft60q8RdGszDo1Dw9CZcZMQJ6XnSOHUBefQxJc4+ilngpY/Kv1uZ3LQ2PmXndLstmM139v5vv6u'
    '18mvWo9PrA9bpbN/fXn70QWt/+7cTKzHf7b1s5mZ1GRyIevDcflXVE08tvTYzba3n20u42oAp0JBPsFPAXFMQLDacaF5M6bDTwiS'
    'U1UnZjZBujI/g54pG7sVq+lO6PsRLMhx/JFVfCeT8albCHh46jj+ALtsnI2HldEatvFymJrtfrLvR9HxGHuJLO3lRzCOw/wIYJgO'
    '5gHGcSxM51+KZ4bG4zDMt5kXmaGcGcpxLB9S2g+m4+fk5vJHmudpmmVYRsvS60GJ5S3L4Ou3hvkGDEwHlF6Xa/xt4xVyvg6wd3qu'
    'QrBI8UrEIsVzDYg/b8DIc//bxnSAgb0FrHZA368DNeXnpCm8Vcw3rINxJM8xBGrRX6NZhmQng4///WBdkqZ57kcA83uQphgC3Ygj'
    'mAfgA4akqT0HX5xH4dM5FR7/G1h8A1BLAwQUAAAACAAqSv9cl4q7HMAAAAATAgAACwAAAF9yZWxzLy5yZWxznZK5bsMwDEB/xdCe'
    'MAfQIYgzZfEWBPkBVqIP2BIFikWdv6/apXGQCxl5PTwS3B5pQO04pLaLqRj9EFJpWtW4AUi2JY9pzpFCrtQsHjWH0kBE22NDsFos'
    'PkAuGWa3vWQWp3OkV4hc152lPdsvT0FvgK86THFCaUhLMw7wzdJ/MvfzDDVF5UojlVsaeNPl/nbgSdGhIlgWmkXJ06IdpX8dx/aQ'
    '0+mvYyK0elvo+XFoVAqO3GMljHFitP41gskP7H4AUEsDBBQAAAAIACpK/1x9cxfuVwEAAFwCAAAPAAAAeGwvd29ya2Jvb2sueG1s'
    'jZFJT8QwDIX/SpU7tDMgltF0LiAWCQFiPWcad2qRxJXjocCvx21VFnHhlLxn6fnly7IjflkTvWRvwcdUmkakXeR5qhoINu1SC1En'
    'NXGwopI3eWoZrEsNgASfz4viIA8Wo1ktp6xbzlfL/vKE0KVvv5fZKyZco0d5L81w92CygBEDfoArTWGy1FB3QYwfFMX6+4rJ+9LM'
    'xsETsGD1x77v+zzYdRqct2eMjrrS7MzmGvj+W3aDekYnTWnmxeHel3cBuGlEI2bFvppi13dWkEpzUKiskZMMi4aathJ8Bd05qq3Q'
    'GXoBPrUC50zbFuOmb6Mw8h80BnLTOWJf8H/AU11jBadUbQNEGckz+L5gTA22yWTRBijNGXH22HqyThv0eHTRpRtRiZb7AZ4XqAO+'
    'dGPNqZuDGiO4a41L6ivu6paz/hhyjmbF/Fh5bL0/Ue8mXumu6anTb68+AVBLAwQUAAAACAAqSv9cJB6boq0AAAD4AQAAGgAAAHhs'
    'L19yZWxzL3dvcmtib29rLnhtbC5yZWxztZE9DoMwDIWvEuUANVCpQwVMXVgrLhAF8yMSEsWuCrcvhQGQOnRhsp4tf+/JTp9oFHdu'
    'oLbzJEZrBspky+zvAKRbtIouzuMwT2oXrOJZhga80r1qEJIoukHYM2Se7pminDz+Q3R13Wl8OP2yOPAPMLxd6KlFZClKFRrkTMJo'
    'tjbBUuLLTJaiqDIZiiqWcFog4skgbWlWfbBPTrTneRc390WuzeMJrt8McHh0/gFQSwMEFAAAAAgAKkr/XGWQeZIZAQAAzwMAABMA'
    'AABbQ29udGVudF9UeXBlc10ueG1srZNNTsMwEIWvEmVbJS4sWKCmG2ALXXABY08aq/6TZ1rS2zNO2kqgEhWFTax43rzPnpes3o8R'
    'sOid9diUHVF8FAJVB05iHSJ4rrQhOUn8mrYiSrWTWxD3y+WDUMETeKooe5Tr1TO0cm+peOl5G03wTZnAYlk8jcLMakoZozVKEtfF'
    'wesflOpEqLlz0GBnIi5YUIqrhFz5HXDqeztASkZDsZGJXqVjleitQDpawHra4soZQ9saBTqoveOWGmMCqbEDIGfr0XQxTSaeMIzP'
    'u9n8wWYKyMpNChE5sQR/x50jyd1VZCNIZKaveCGy9ez7QU5bg76RzeP9DGk35IFiWObP+HvGF/8bzvERwu6/P7G81k4af+aL4T9e'
    'fwFQSwECFAMUAAAACAAqSv9cRsdNSJUAAADNAAAAEAAAAAAAAAAAAAAAgAEAAAAAZG9jUHJvcHMvYXBwLnhtbFBLAQIUAxQAAAAI'
    'ACpK/1ypO90dEQEAAGYCAAARAAAAAAAAAAAAAACAAcMAAABkb2NQcm9wcy9jb3JlLnhtbFBLAQIUAxQAAAAIACpK/1zBFxC+kgYA'
    'AMYgAAATAAAAAAAAAAAAAACAAQMCAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAhQDFAAAAAgAKkr/XKQpsLjxAgAAKQoAABgAAAAA'
    'AAAAAAAAAICBxggAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQIUAxQAAAAIACpK/1yMbVtt7wIAACgQAAANAAAAAAAAAAAA'
    'AACAAe0LAAB4bC9zdHlsZXMueG1sUEsBAhQDFAAAAAgAKkr/XJeKuxzAAAAAEwIAAAsAAAAAAAAAAAAAAIABBw8AAF9yZWxzLy5y'
    'ZWxzUEsBAhQDFAAAAAgAKkr/XH1zF+5XAQAAXAIAAA8AAAAAAAAAAAAAAIAB8A8AAHhsL3dvcmtib29rLnhtbFBLAQIUAxQAAAAI'
    'ACpK/1wkHpuirQAAAPgBAAAaAAAAAAAAAAAAAACAAXQRAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQIUAxQAAAAIACpK'
    '/1xlkHmSGQEAAM8DAAATAAAAAAAAAAAAAACAAVkSAABbQ29udGVudF9UeXBlc10ueG1sUEsFBgAAAAAJAAkAPgIAAKMTAAAAAA=='
)
PHILIPPINE_ZONE = ZoneInfo("Asia/Manila")


class WarehouseConversionError(ValueError):
    """Raised when an uploaded SAP workbook does not match the Warehouse format."""


@dataclass(frozen=True)
class WarehouseFilenameMetadata:
    source_stem: str
    location: str
    remarks: str
    company_code: str
    sap_no: str
    stock_status: str


@dataclass(frozen=True)
class WarehouseSourceRecord:
    item_no: str
    product: str
    uom: Any
    record_qty: Any


@dataclass(frozen=True)
class WarehouseConversionResult:
    output_bytes: bytes
    output_filename: str
    metadata: WarehouseFilenameMetadata
    process_date: date
    auditor_name: str
    records: tuple[WarehouseSourceRecord, ...]
    source_signature: str

    @property
    def row_count(self) -> int:
        return len(self.records)

    def preview_rows(self, limit: int = 200) -> list[dict[str, Any]]:
        shown = self.records[: max(0, int(limit))]
        return [
            {
                "login_date": self.process_date.isoformat(),
                "sap_no": self.metadata.sap_no,
                "sap_date": self.process_date.isoformat(),
                "locations": self.metadata.location,
                "stock_status": self.metadata.stock_status,
                "task": "Balance",
                "product": record.product,
                "uom": record.uom,
                "record_qty": record.record_qty,
                "count_qty": None,
                "remarks": self.metadata.remarks,
                "auditor_name": self.auditor_name,
            }
            for record in shown
        ]


def philippine_today() -> date:
    return datetime.now(PHILIPPINE_ZONE).date()


def _clean_uploaded_stem(filename: str) -> str:
    raw_name = Path(str(filename or "").strip()).name
    if not raw_name:
        raise WarehouseConversionError("The uploaded Excel filename is missing.")
    stem = Path(raw_name).stem
    # Browsers commonly append duplicate-download suffixes such as (1).
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem).strip()
    stem = re.sub(r"\s+", " ", stem)
    if not stem:
        raise WarehouseConversionError("The uploaded Excel filename is invalid.")
    return stem


def parse_warehouse_filename(
    filename: str,
    *,
    process_date: date | None = None,
) -> WarehouseFilenameMetadata:
    conversion_date = process_date or philippine_today()
    stem = _clean_uploaded_stem(filename)
    words = stem.split()
    if len(words) < 3:
        raise WarehouseConversionError(
            "Filename must follow the Warehouse pattern, for example: "
            "Cebu Damage Warehouse EPLSI.xlsx or Cebu Good Stocks EPLSI.xlsx."
        )

    location = words[0].strip()
    remarks_word = words[1].strip()
    company_code = words[-1].strip().upper()
    location_code = re.sub(r"[^A-Za-z0-9]", "", location)[:3].upper()

    if len(location_code) < 3:
        raise WarehouseConversionError(
            "The first filename word must contain at least three letters for sap_no."
        )
    if not company_code:
        raise WarehouseConversionError("The last filename word/company code is missing.")

    remarks_key = remarks_word.casefold()
    if remarks_key == "damage":
        remarks = "Damage"
        stock_suffix = "DW"
    elif remarks_key == "good":
        remarks = "Good"
        stock_suffix = "GS"
    else:
        raise WarehouseConversionError(
            "The second filename word must be either Damage or Good."
        )

    sap_no = f"{location_code}{conversion_date:%y%m%d}-{company_code}"
    stock_status = f"{company_code}-{stock_suffix}"
    return WarehouseFilenameMetadata(
        source_stem=stem,
        location=location,
        remarks=remarks,
        company_code=company_code,
        sap_no=sap_no,
        stock_status=stock_status,
    )


def _normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def _remove_apostrophes(value: Any) -> str:
    text = "" if value is None else str(value)
    for apostrophe in ("'", "’", "‘", "ʼ", "`"):
        text = text.replace(apostrophe, "")
    return text.strip()


def _find_header_row_and_columns(worksheet: Any) -> tuple[int, dict[str, int]]:
    expected = {
        "item no": "item_no",
        "item description": "product",
        "inventory uom": "uom",
        "in stock": "record_qty",
    }
    max_scan_rows = min(max(worksheet.max_row, 1), 15)
    max_scan_cols = min(max(worksheet.max_column, 1), 30)

    for row_index in range(1, max_scan_rows + 1):
        found: dict[str, int] = {}
        for column_index in range(1, max_scan_cols + 1):
            normalized = _normalize_header(worksheet.cell(row_index, column_index).value)
            field_name = expected.get(normalized)
            if field_name:
                found[field_name] = column_index
        if set(found) == set(expected.values()):
            return row_index, found

    raise WarehouseConversionError(
        "Required SAP columns were not found: Item No., Item Description, "
        "Inventory UoM, and In Stock."
    )


def extract_warehouse_records(excel_bytes: bytes) -> tuple[WarehouseSourceRecord, ...]:
    if not excel_bytes:
        raise WarehouseConversionError("The uploaded SAP Excel file is empty.")

    try:
        workbook = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise WarehouseConversionError(
            "The uploaded file could not be opened as a valid .xlsx workbook."
        ) from exc

    worksheet = workbook.active
    header_row, columns = _find_header_row_and_columns(worksheet)

    start_row: int | None = None
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        item_no = worksheet.cell(row_index, columns["item_no"]).value
        if str(item_no or "").strip().upper() == "A01AMB01":
            start_row = row_index
            break

    if start_row is None:
        raise WarehouseConversionError(
            'Starting item "A01AMB01" was not found under the Item No. column.'
        )

    records: list[WarehouseSourceRecord] = []
    for row_index in range(start_row, worksheet.max_row + 1):
        item_no_value = worksheet.cell(row_index, columns["item_no"]).value
        product_value = worksheet.cell(row_index, columns["product"]).value
        uom_value = worksheet.cell(row_index, columns["uom"]).value
        quantity_value = worksheet.cell(row_index, columns["record_qty"]).value

        if all(value in (None, "") for value in (item_no_value, product_value, uom_value, quantity_value)):
            break

        item_no_text = str(item_no_value or "").strip()
        product_raw = "" if product_value is None else str(product_value).strip()
        total_marker = f"{item_no_text} {product_raw}".casefold()
        if "grand total" in total_marker or total_marker.strip() == "total" or item_no_text.casefold() == "total":
            break

        if not product_raw:
            raise WarehouseConversionError(
                f"Item Description is blank at source row {row_index}."
            )

        product = _remove_apostrophes(product_raw)
        uom = None if uom_value in (None, "") else uom_value
        records.append(
            WarehouseSourceRecord(
                item_no=item_no_text,
                product=product,
                uom=uom,
                record_qty=quantity_value,
            )
        )

    if not records:
        raise WarehouseConversionError("No Warehouse product rows were captured.")

    return tuple(records)


def _safe_output_filename(metadata: WarehouseFilenameMetadata, auditor_name: str) -> str:
    first_name = str(auditor_name or "Auditor").strip().split()[0] or "Auditor"
    raw = f"For Upload {metadata.company_code} {metadata.remarks} - {first_name}.xlsx"
    return re.sub(r'[<>:"/\\|?*]+', "_", raw)


def _load_template(template_path: Path | None = None):
    path = Path(template_path or WAREHOUSE_TEMPLATE_PATH)
    try:
        if path.exists():
            workbook = load_workbook(path)
        else:
            # Self-contained fallback: use the embedded approved workbook.
            template_bytes = base64.b64decode(WAREHOUSE_TEMPLATE_XLSX_BASE64)
            workbook = load_workbook(BytesIO(template_bytes))
    except Exception as exc:
        raise WarehouseConversionError(
            "The Warehouse conversion template could not be opened."
        ) from exc
    worksheet = workbook["For Uploading"] if "For Uploading" in workbook.sheetnames else workbook.active
    headers = [worksheet.cell(1, column).value for column in range(1, 14)]
    if headers != WAREHOUSE_OUTPUT_HEADERS:
        raise WarehouseConversionError(
            "The Warehouse conversion template headers do not match the approved format."
        )
    return workbook, worksheet


def _write_output_rows(
    worksheet: Any,
    records: Iterable[WarehouseSourceRecord],
    metadata: WarehouseFilenameMetadata,
    process_date: date,
    auditor_name: str,
) -> int:
    records = tuple(records)
    style_prototypes = [copy(worksheet.cell(2, column)._style) for column in range(1, 14)]
    alignment_prototypes = [copy(worksheet.cell(2, column).alignment) for column in range(1, 14)]
    protection_prototypes = [copy(worksheet.cell(2, column).protection) for column in range(1, 14)]

    clear_through = max(worksheet.max_row, len(records) + 1)
    for row_index in range(2, clear_through + 1):
        for column_index in range(1, 14):
            worksheet.cell(row_index, column_index).value = None

    excel_date = datetime.combine(process_date, datetime.min.time())
    for offset, record in enumerate(records, start=2):
        values = [
            None,
            excel_date,
            metadata.sap_no,
            excel_date,
            metadata.location,
            metadata.stock_status,
            "Balance",
            record.product,
            record.uom,
            record.record_qty,
            None,
            metadata.remarks,
            auditor_name,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(offset, column_index)
            cell._style = copy(style_prototypes[column_index - 1])
            cell.alignment = copy(alignment_prototypes[column_index - 1])
            cell.protection = copy(protection_prototypes[column_index - 1])
            cell.value = value
            cell.number_format = "yyyy-mm-dd" if column_index in (2, 4) else "General"

    target_last_row = len(records) + 1
    if worksheet.max_row > target_last_row:
        worksheet.delete_rows(
            target_last_row + 1,
            worksheet.max_row - target_last_row,
        )
    return len(records)


def build_warehouse_conversion(
    excel_bytes: bytes,
    filename: str,
    auditor_name: str,
    *,
    process_date: date | None = None,
    template_path: Path | None = None,
) -> WarehouseConversionResult:
    conversion_date = process_date or philippine_today()
    clean_auditor_name = str(auditor_name or "").strip()
    if not clean_auditor_name:
        raise WarehouseConversionError(
            "The signed-in user's full name is required for auditor_name."
        )

    metadata = parse_warehouse_filename(filename, process_date=conversion_date)
    records = extract_warehouse_records(excel_bytes)
    workbook, worksheet = _load_template(template_path)
    _write_output_rows(worksheet, records, metadata, conversion_date, clean_auditor_name)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    buffer = BytesIO()
    workbook.save(buffer)
    output_bytes = buffer.getvalue()

    # Re-open the produced workbook as an integrity check before exposing it.
    try:
        verification_book = load_workbook(BytesIO(output_bytes), data_only=False, read_only=True)
        verification_sheet = verification_book["For Uploading"]
        if verification_sheet.max_row != len(records) + 1:
            raise WarehouseConversionError(
                "Converted row count did not match the captured SAP product count."
            )
    except WarehouseConversionError:
        raise
    except Exception as exc:
        raise WarehouseConversionError(
            "The converted Excel file failed the final workbook integrity check."
        ) from exc

    return WarehouseConversionResult(
        output_bytes=output_bytes,
        output_filename=_safe_output_filename(metadata, clean_auditor_name),
        metadata=metadata,
        process_date=conversion_date,
        auditor_name=clean_auditor_name,
        records=records,
        source_signature=hashlib.sha256(excel_bytes).hexdigest(),
    )


def render_warehouse_conversion_page(user: dict[str, Any]) -> None:
    import pandas as pd
    import streamlit as st

    st.markdown(
        """
        <style>
        .iars-excel-hero {
            border: 1px solid #DDE5EF;
            border-radius: 16px;
            padding: 1.05rem 1.15rem;
            margin: 0 0 .9rem 0;
            background: linear-gradient(135deg, #F8FAFD 0%, #FFFFFF 58%, #FFF9EB 100%);
            box-shadow: 0 8px 24px rgba(6,26,54,.06);
        }
        .iars-excel-hero h2 { margin: 0; color: #061A36; font-size: 1.35rem; }
        .iars-excel-hero p { margin: .28rem 0 0; color: #667085; font-size: .88rem; }
        .iars-excel-route {
            display: grid;
            grid-template-columns: 1fr auto 1fr auto 1fr;
            align-items: center;
            gap: .55rem;
            margin-top: .85rem;
        }
        .iars-excel-route div {
            min-height: 64px;
            border: 1px solid #E4EAF2;
            border-radius: 12px;
            padding: .65rem .72rem;
            background: rgba(255,255,255,.9);
        }
        .iars-excel-route strong { display:block; color:#061A36; font-size:.86rem; }
        .iars-excel-route span { color:#667085; font-size:.74rem; line-height:1.25; }
        .iars-excel-route b { color:#C78B12; font-size:1.1rem; }
        @media (max-width: 760px) {
            .iars-excel-route { grid-template-columns: 1fr; }
            .iars-excel-route b { display:none; }
        }
        </style>
        <div class="iars-excel-hero">
          <h2>Warehouse Excel Conversion</h2>
          <p>Convert an SAP Warehouse stock export into the approved upload template without changing the original file.</p>
          <div class="iars-excel-route">
            <div><strong>1. SAP Excel</strong><span>Upload the Warehouse stock file.</span></div>
            <b>→</b>
            <div><strong>2. IARS Conversion</strong><span>Validate, map and clean the required data.</span></div>
            <b>→</b>
            <div><strong>3. Compatible Output</strong><span>Download the approved 13-column template.</span></div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        st.markdown("### Upload SAP Warehouse File")
        st.caption(
            "Accepted filename examples: Cebu Damage Warehouse EPLSI.xlsx or "
            "Cebu Good Stocks EPLSI.xlsx"
        )
        uploaded_file = st.file_uploader(
            "SAP Warehouse Excel",
            type=["xlsx"],
            key="warehouse_sap_excel_uploader_v4_5_12",
            help="The original SAP file remains unchanged.",
        )

    if uploaded_file is None:
        st.info(
            "Upload an SAP Warehouse .xlsx file. IARS will capture Item Description, "
            "Inventory UoM and In Stock beginning at item A01AMB01."
        )
        return

    auditor_name = str(
        user.get("full_name") or user.get("username") or ""
    ).strip()
    process_date = philippine_today()

    try:
        with st.spinner("Validating and converting the SAP Warehouse file…"):
            result = build_warehouse_conversion(
                uploaded_file.getvalue(),
                uploaded_file.name,
                auditor_name,
                process_date=process_date,
            )
    except WarehouseConversionError as exc:
        st.error(str(exc))
        return
    except Exception as exc:
        st.error(f"Warehouse conversion failed: {exc}")
        return

    metadata = result.metadata
    metric_columns = st.columns(4)
    metric_columns[0].metric("Products Captured", f"{result.row_count:,}")
    metric_columns[1].metric("SAP No.", metadata.sap_no)
    metric_columns[2].metric("Location", metadata.location)
    metric_columns[3].metric("Stock Status", metadata.stock_status)

    st.success(
        f"Conversion completed for {result.row_count:,} product rows. "
        "Apostrophes were removed and the output passed the workbook integrity check."
    )

    with st.expander("Conversion Details", expanded=True):
        details = pd.DataFrame(
            [
                ["Processing Date", result.process_date.isoformat()],
                ["Source Filename", uploaded_file.name],
                ["SAP No.", metadata.sap_no],
                ["Location", metadata.location],
                ["Remarks", metadata.remarks],
                ["Stock Status", metadata.stock_status],
                ["Task", "Balance"],
                ["Auditor Name", result.auditor_name],
            ],
            columns=["Field", "Generated Value"],
        )
        st.dataframe(details, hide_index=True, width="stretch")

    st.markdown("### Converted Data Preview")
    preview = pd.DataFrame(result.preview_rows(limit=200))
    st.dataframe(preview, hide_index=True, width="stretch", height=390)
    if result.row_count > len(preview):
        st.caption(f"Showing the first {len(preview):,} of {result.row_count:,} converted rows.")

    st.download_button(
        "⬇️ Download Converted Warehouse Excel",
        data=result.output_bytes,
        file_name=result.output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"warehouse_download_{result.source_signature[:16]}",
        type="primary",
        width="stretch",
    )
    st.caption(
        "Output format: For Uploading sheet · 13 approved columns · dates in yyyy-mm-dd · "
        "count_qty and id remain blank."
    )

# ---------------------------------------------------------------------------
# Sales Personnel - LOGP Excel Conversion
# ---------------------------------------------------------------------------

LOGP_OUTPUT_HEADERS = [
    "trans_id",
    "login_date",
    "logp_no",
    "employee_name",
    "docu_date",
    "docu_name",
    "sold_no",
    "inv_no",
    "pd_no",
    "prod_code",
    "prod_name",
    "prod_uom",
    "inv_qty",
    "disc_qty",
    "record_qty",
    "count_qty",
    "remarks",
    "auditor_name",
]

LOGP_TEMPLATE_PATH = (
    Path(__file__).resolve().parent / "assets" / "logp_conversion_template.xlsx"
)
# Embedded approved template fallback. The LOGP conversion remains operational
# even when deployment tools omit the assets folder.
LOGP_TEMPLATE_XLSX_BASE64 = (
    'UEsDBBQAAAAIALE+A12YqOmlvQAAACQBAAAPAAAAeGwvd29ya2Jvb2sueG1sjY+7bgMhEEV/BU2fBRLntVrWjZu0yRdgGLzIwKwY'
    'bPP5kZ0obtMdneLq3GnbcxJnrBypGNCDAoHFkY/lYODUwsMbbOepjxeqxz3RUfScCo/dwNLaOkrJbsFseaAVS88pUM228UD1IHmt'
    'aD0viC0n+ajUi8w2Frju3Sz/kSg2o4GvK2sQN/fhDWgQdYzewKdWT+r51e91cHaD9h1+S+p/SiiE6HBH7pSxtJ+Uism2SIWXuDII'
    'OU/yniXvj+dvUEsDBBQAAAAIALE+A13quNcCGgIAAPgHAAANAAAAeGwvc3R5bGVzLnhtbKWVS2+jMBCA/4rlewKk3aqK6lS7kZD2'
    '0kv3sIdeCAxgaWwj41Swv37lB49UagopF4Zh5puHPfbTcyeQvINuuZKMJtuYEpC5KrisGD2bcvNInw9P3b41PcJrDWBIJ1C2+47R'
    '2phmH0VtXoPI2q1qQHYCS6VFZtqt0lXUNhqyorVuAqNdHD9EIuOSWqI8i1SYluTqLA2jyUxJ/Ot3wWjycE+JRx5VAYz2fd+/bYR4'
    '2xQFJdHhKRpRFlAqOTHv6KByJfwj7xkymiTOr9vnCpUmpgYBNr5XykyAtztmyE+aD0EGzKe42z19Iro6MZqmsXuWMYPgK+eIY+X3'
    'vnKOaN9NZgxomXJEEuQ/fQOMSiVhJAbjL50qnfXJ7sdqv1YhL3xe1XFesX1CxdGF/3f5YWm/QAfBNfGkdAF6bOOOTkq/IF62EkJp'
    'iBsLRk0dNrVfSS4L6KBg1G5eH9paWwPNq3qFmzO3FkY1y72Manyuxiix3M3bB9GXOYquOzkgvlra3/LD3HblbGZjO7FyFDliED0q'
    'fPiYc+QQYkZ/vJXelVOY9YBkBsiaBvtf7td0SnwGTCbg7iowVb554YujPQ9WBosXB1uOD+ft8v465MtZnECn7phelffdVeyY91Xc'
    'bmmWlz3/ibySAqY9nA0Kex8anttTNwdpQA+btStvKWvVcoSRmE2Dm44P4zbqib0gGH2xncfLXT8frtZ9Tnf44T9QSwMEFAAAAAgA'
    'sT4DXVhXW6vGAgAAhAoAABMAAAB4bC90aGVtZS90aGVtZTEueG1svVZbb5swGP0rlt9XINzaqLRqSdAeNk1aJu3ZwQa8GBvZTpP+'
    '+wlzD6HqNnXkIb6c853zffjC/eO5ZOCFSEUFj6BzY0NAeCow5XkEjzr7dAsfH+7RWhekJICjkkTwW5bRlEBwLhlXaxTBQutqbVkq'
    'LUiJ1I2oCD+XLBOyRFrdCJlbWKIT5XnJrJVtB1aJKId92C0jJeFa1QMpk7t0rlVP4YNT/6lXFTMJXhCL4IlyLE4/yFlDwJDSMZMR'
    'tM0DgfVwb/UsphfII2Jino7YMvBhZYgy3/dMz/O94GlQMAim58BtuA22wRDRIFCaEt7amUYNV7HXgUeopnkl+ibcuM6UMFJwZ4Qn'
    'v/5NCQbVNL0ZIUniUSlHqKbpzwj+893z5kLBoJpmMCOE9tPGC6cEgyoY5YcZ3PYDN+5T7jGZYJ+v4u98LwlXHX6AWaOV1gTgemnd'
    'leiXkIng2rxlpCkH+rUiGUpJBGPE6F5So4DWBC1OpWphyrpQKCn/aLlBwRqnbgpRLtYho4zt9CsjX5TxpgSjOKGMmY4h9XWvipjJ'
    'Tm8KfIOFD84fc9p9coVnzS0zPu2BUwQD17fh3+dTSaU3SBUNzkz1+50PKs4qtP+HzJ39kdlYlyUkWUZSvTAydL8o3Ua5Ov2v6Loj'
    'jprIXYFPYM+O8jvCEfRDx7chwFTprjIAUxlB32uvCMRyHsFUSwik0D+pLnYFqkgE2xpOjp6GYw4zVhWoGQ3c0QHZ4k279zNKxFi9'
    'TGvab7PZ58mH77aGZW5hynV3XnbpoLVC+qvAzbgTju/UPtBcMJdoaKvWe65AJVRf1XfY6KuK1qpAmLTDt8MwO5aDO3u1YNtftp2r'
    'sbUG+F57t9ft2Uv23AV77nvtOcPiW/Q3cjKsygvBuk5vCJq/7nSkHKD6K7DfLSpFjOD6NbYBhpdtzZdsffx2l4npXXzrdSMPvwFQ'
    'SwMEFAAAAAgAsT4DXYeonh/pAAAAPAQAABQAAAB4bC9zaGFyZWRTdHJpbmdzLnhtbJ3RQW7EIAwF0KuM2DekXVRVlGSOEiHwJKhg'
    'U+xEzO2rpNuqkrsxG54/NuO95XQ7oHIknMxr15sboKcQcZ3MLo+XD3OfxzYwy63lhDy0yWwiZbCW/QbZcUcFsOX0oJqdcEd1tVwq'
    'uMAbgORk3/r+3WYX0Vyt4lmvdgMX52EypQJDPcDMUh3yEsNo2yDzWX+u/4kSrRGX4AS0rCxIKgO5JHoCLOiyLi2Q3/VvvJQ6iykF'
    '7WQRDy0p6pBSKSyeAuiVegmX2imrt/AlT90fRfZqVMFTDWrmaUf5R1h29ZNVxu0hCtVf9n4eLPM3UEsDBBQAAAAIALE+A13oDssA'
    'BggAACU0AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sjdtdc9pGFMbxr8JwXxtJvGbqdNpKq9X7am96TWP8MjXGAyTxx+8I'
    'hL1/7Uniq/AjzwHzEI8Pa+X3P163T6Nvm/3hcfd8Mw6uJuPR5vnL7vbx+f5m/PV499ty/Mfn318/fd/t/zs8bDbH0ev26fnw6fVm'
    '/HA8vny6vj58edhs14er3cvm+XX7dLfbb9fHw9Vuf399eNlv1rense3TdTiZzK+368fncfeAp3vVKWz2o9vN3frr09HuvuvN4/3D'
    '8WYczMaj6y74Zfd06P8cbR+7L3I82q5fT39+f7w9PtyMF1fhchbM5uFsPPry9XDcbf85/0Xw/hjn2bCfDd9mg/BqGs4Wy+DXw1E/'
    'HL0PBx9+5mk/PH0bDldXy9lsOl8ufjk864dn788cXS2CySr69ey8n527L3m2mEQfeMXBW9fvZYerD7/m4FJ3d+P9yT88fnnZ3Y3L'
    '+OTj45dX3t14f8M++tKXl+nl+/Ty40++uoyv3K/9Z09+/f5v/fTNEa+P6w773ffRvgudnqC7+WcwHh1OY8eb8eF0/7fPk+4Bvp0f'
    '5i351zkZIRlIyb/PyRDJUErG0mNGUjKRklMpqaTkTEqm5+QCybmU1FJyISUzKbmUkrmUXEnJQmxefJNKMSq+S5X0/IH4NtViVHyf'
    'GvELEN8oIz6q+E614qOKb5UVo3yvrk/fCc43ROh8Q4SX77/TP/rw9GhT976/kYihRMgrJFJIQxmUQwVUQhVUQw1koBayZ3nlRE45'
    'EcqJhHKQiKFEyCskUkhDGZRDBVRCFVRDDWSgFrKRXM7UKQcv7K+pUA4UQ4mQV1AKaSiDcqiASqiCaqiBDNRCdiqXM3PKmaGcmVAO'
    'EjGUCHmFRAppKINyqIBKqIJqqIEM1EJ2Jpczd8qZo5y5UA4SMZQIeYVECmkog3KogEqogmqogQzUQnYul7NwylmgnIVQDhIxlAh5'
    'hUQKaSiDcqiASqiCaqiBDNRCdiGXs3TKWaKcpVAOEjGUCHmFRAppKINyqIBKqIJqqIEM1EJ2KZezcspZoZyVUA4SMZQIeYVECmko'
    'g3KogEqogmqogQzUQnYllxNM3M1/gno6ev0wE5OJNKKYSUlNZmROFmRJVmRNNqQhW9L29PvCJ6WAfQVSX8jEZCKNKGZSUpMZmZMF'
    'WZIVWZMNaciWtD39vtxFuvuk7vYlrdLMxGQijShmUlKTGZmTBVmSFVmTDWnIlrQ9/b7c3Trgct3R74vrNZlII4qZlNRkRuZkQZZk'
    'RdZkQxqyJW1Pvy933Q64b3f0++LGTSbSiGImJTWZkTlZkCVZkTXZkIZsSdvT78vdwN+PQc99STs4MzGZSCOKmZTUZEbmZEGWZEXW'
    'ZEMasiVtT78vdynvzvrcvqS1nJmYTKQRxUxKajIjc7IgS7Iia7IhDdmStqffl7unB1zUO/p9cVUnE2lEMZOSmszInCzIkqzImmxI'
    'Q7ak7en35a7u3Wmw25e0vDMTk4k0ophJSU1mZE4WZElWZE02pCFb0vb0+3K3+e742+1L2ueZiclEGlHMpKQmMzInC7IkK7ImG9KQ'
    'LWl7+ueY7n4fcr/v6B9lcr8nE2lEMZOSmszInCzIkqzImmxIQ7ak7en35e73Iff7jn5f3O/JRBpRzKSkJjMyJwuyJCuyJhvSkC1p'
    'e/p94aB8cFIuHpUPzsoHh+XiafnguHxwXj44MB+cmA+OzAdn5oND88Gp+eDYfHBuPjg4H5yc/2C/D939PuR+39Hvi/s9mUgjipmU'
    '1GRG5mRBlmRF1mRDGrIlbU+/L3e/D7nfd/T74n5PJtKIYiYlNZmROVmQJVmRNdmQhmxJ29Pvy93vu98Fu31J+z0zMZlII4qZlNRk'
    'RuZkQZZkRdZkQxqyJW1Pvy93vw+533f0++J+TybSiGImJTWZkTlZkCVZkTXZkIZsSdvT78vd70Pu9x39vrjfk4k0ophJSU1mZE4W'
    'ZElWZE02pCFb0vb0+3L3+5D7fUe/L+73ZCKNKGZSUpMZmZMFWZIVWZMNaciWtD39vtz9PuR+39Hvi/s9mUgjipmU1GRG5mRBlmRF'
    '1mRDGrIlbU//V/Hufh9xv+/o/zae+z2ZSCOKmZTUZEbmZEGWZEXWZEMasiVtT78vd7+PuN939Pvifk8m0ohiJiU1mZE5WZAlWZE1'
    '2ZCGbEnb0+/L3e8j7vcd/b6435OJNKKYSUlNZmROFmRJVmRNNqQhW9L29PvCtTGDi2PEq2MGl8cMro8RL5AZXCEzuERmcI3M4CKZ'
    'wVUyg8tkBtfJDC6UGVwpM7hUZnCtzOBimR/s95G730fc7zv6fXG/JxNpRDGTkprMyJwsyJKs+mfn1ird2XDQkC1pe/rdubt+xF2/'
    'o98dd30ykUYUMympyYzMyYIsyap/9kF3wp0NBw3Zkran352790fc+zv63XHvJxNpRDGTkprMyJwsyJKs+mcfdCfc2XDQkC1pe/rd'
    'uZ8BIn4G6Oh3x88AZCKNKGZSUpMZmZMFWZIVWZMNaciWtD39vtzPABE/A3T0++JnADKRRhQzKanJjMzJgizJiqzJhjRkS9qefl/O'
    'Z4A4Om/wg283d633r8B0duJ4OhHm+zt/NO/siPE0kOaDn847O1M8DaX58KfzkTt/3gAG8+5a4M87P1Pj6fnHz2De/TF5mb8e/NeA'
    'l/X9plrv7x+fD6Onzd3xZjy5WoxH+/N/oTndPu5eTrdm49G/u+Nxt73oYbO+3ew7RePR3W53fMP5Cd/+x8/n/wFQSwMEFAAAAAAA'
    'sT4DXWWaBBooAQAAKAEAAAsAAABfcmVscy8ucmVsc++7vzw/eG1sIHZlcnNpb249IjEuMCIgZW5jb2Rpbmc9InV0Zi04Ij8+PFJl'
    'bGF0aW9uc2hpcHMgeG1sbnM9Imh0dHA6Ly9zY2hlbWFzLm9wZW54bWxmb3JtYXRzLm9yZy9wYWNrYWdlLzIwMDYvcmVsYXRpb25z'
    'aGlwcyI+PFJlbGF0aW9uc2hpcCBUeXBlPSJodHRwOi8vc2NoZW1hcy5vcGVueG1sZm9ybWF0cy5vcmcvb2ZmaWNlRG9jdW1lbnQv'
    'MjAwNi9yZWxhdGlvbnNoaXBzL29mZmljZURvY3VtZW50IiBUYXJnZXQ9Ii94bC93b3JrYm9vay54bWwiIElkPSJSYWNkZDUyNGEx'
    'NjkyNGE0NyIgLz48L1JlbGF0aW9uc2hpcHM+UEsDBBQAAAAIALE+A10relt+EAEAAPICAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2su'
    'eG1sLnJlbHO1kktOwzAURbdieU6cX9MGNe2ECdPSDTj2dRLVn8h2IV0bA5bEFhAUoQQxYNLJG9wnHZ139d5f37b7yWjyDB8GZxua'
    'JSklsMLJwXYNPUd1t6H73fYAzePgbOiHMZDJaBsa2sc43jMWRA/DQ+JG2Mlo5bzhMSTOd2zk4sQ7sDxNK+bnDLpkkuNlxH+ITqlB'
    '4MGJs4GNf4BZiBeNQMmR+w6xoWzS31kyGU3Jo2zoQbUlCrTlGpuirCAoYTcTij0Mlj5f0XVmMysukaPOIEVdlbytb2kVeu4hn6If'
    'bPe7rflqplcoyBXP61Wh0rJap7fUe3H+FHogLtV+4s8DgDhvL0uLdLWWbaYEL8Gv7bHF5+4+AFBLAwQUAAAACACxPgNdjYLZqRYB'
    'AABTAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk0FOwzAQRa8SeYtqpywQQkm7ALaABBewnEli1R5bnmlIz8aCI3EFVAdFgJAi'
    '1G48m/F7/y/m4+292o7eFQMksgFrsZalKABNaCx2tdhzu7oW2031cohAxegdUi165nijFJkevCYZIuDoXRuS10wypE5FbXa6A3VZ'
    'llfKBGRAXvGRITbVHbR677i4Hxlw0o7eieJ22juqaqFjdNZotgHVgM0vySq0rTXQBLP3gCwpJtAN9QDsncxTem3xIoPVn84Ejv4n'
    '/WolE7i8Q72NNCseB0jJNlA86cQP2kMt1OgU8cEByTM3zNAlNffgYXrXJwfImMWyvU7QPHOy2J2983f2UpDXkHb5I6k8Tu//M8zM'
    'n4OofCKbT1BLAQIUAxQAAAAIALE+A12YqOmlvQAAACQBAAAPAAAAAAAAAAAAAACkgQAAAAB4bC93b3JrYm9vay54bWxQSwECFAMU'
    'AAAACACxPgNd6rjXAhoCAAD4BwAADQAAAAAAAAAAAAAApIHqAAAAeGwvc3R5bGVzLnhtbFBLAQIUAxQAAAAIALE+A11YV1urxgIA'
    'AIQKAAATAAAAAAAAAAAAAACkgS8DAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAhQDFAAAAAgAsT4DXYeonh/pAAAAPAQAABQAAAAA'
    'AAAAAAAAAKSBJgYAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAhQDFAAAAAgAsT4DXegOywAGCAAAJTQAABgAAAAAAAAAAAAAAKSB'
    'QQcAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQIUAxQAAAAAALE+A11lmgQaKAEAACgBAAALAAAAAAAAAAAAAACkgX0PAABf'
    'cmVscy8ucmVsc1BLAQIUAxQAAAAIALE+A10relt+EAEAAPICAAAaAAAAAAAAAAAAAACkgc4QAAB4bC9fcmVscy93b3JrYm9vay54'
    'bWwucmVsc1BLAQIUAxQAAAAIALE+A12NgtmpFgEAAFMDAAATAAAAAAAAAAAAAACkgRYSAABbQ29udGVudF9UeXBlc10ueG1sUEsF'
    'BgAAAAAIAAgAAwIAAF0TAAAAAA=='
)


class LogpConversionError(ValueError):
    """Raised when an uploaded workbook does not match the approved LOGP format."""


@dataclass(frozen=True)
class LogpFilenameMetadata:
    source_stem: str
    logp_no: str
    docu_date: date
    docu_name: str
    remarks: str


@dataclass(frozen=True)
class LogpSourceRecord:
    item_no: str
    product_name: str
    prod_uom: Any
    record_qty: Any
    warehouse_asr: str
    employee_name: str


@dataclass(frozen=True)
class LogpConversionResult:
    output_bytes: bytes
    output_filename: str
    metadata: LogpFilenameMetadata
    process_date: date
    auditor_name: str
    records: tuple[LogpSourceRecord, ...]
    source_signature: str

    @property
    def row_count(self) -> int:
        return len(self.records)

    def preview_rows(self, limit: int = 200) -> list[dict[str, Any]]:
        shown = self.records[: max(0, int(limit))]
        return [
            {
                "trans_id": None,
                "login_date": self.process_date.isoformat(),
                "logp_no": self.metadata.logp_no,
                "employee_name": record.employee_name,
                "docu_date": self.metadata.docu_date.isoformat(),
                "docu_name": self.metadata.docu_name,
                "sold_no": None,
                "inv_no": None,
                "pd_no": None,
                "prod_code": None,
                "prod_name": record.product_name,
                "prod_uom": record.prod_uom,
                "inv_qty": None,
                "disc_qty": None,
                "record_qty": record.record_qty,
                "count_qty": None,
                "remarks": self.metadata.remarks,
                "auditor_name": self.auditor_name,
            }
            for record in shown
        ]


def _clean_logp_uploaded_stem(filename: str) -> str:
    raw_name = Path(str(filename or "").strip()).name
    if not raw_name:
        raise LogpConversionError("The uploaded LOGP Excel filename is missing.")
    stem = Path(raw_name).stem
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem).strip()
    stem = re.sub(r"\s+", " ", stem)
    if not stem:
        raise LogpConversionError("The uploaded LOGP Excel filename is invalid.")
    return stem


def parse_logp_filename(filename: str) -> LogpFilenameMetadata:
    stem = _clean_logp_uploaded_stem(filename)

    docu_name_match = re.match(r"\s*(LOGP)\b", stem, flags=re.I)
    if not docu_name_match:
        raise LogpConversionError(
            "The filename must begin with LOGP, for example: "
            "LOGP 61005123 DAVIDO GOOD - 06-22-2026.xlsx."
        )
    docu_name = docu_name_match.group(1).upper()

    number_match = re.search(r"\bLOGP[\s_-]+(\d+)\b", stem, flags=re.I)
    if not number_match:
        raise LogpConversionError(
            "The LOGP number was not found after the word LOGP in the filename."
        )
    logp_no = number_match.group(1)

    date_matches = list(
        re.finditer(
            r"(?<!\d)(0?[1-9]|1[0-2])[-_/](0?[1-9]|[12]\d|3[01])[-_/](\d{4})(?!\d)",
            stem,
        )
    )
    if not date_matches:
        raise LogpConversionError(
            "The document date was not found in the filename. Use MM-DD-YYYY."
        )
    month, day, year = map(int, date_matches[-1].groups())
    try:
        docu_date = date(year, month, day)
    except ValueError as exc:
        raise LogpConversionError("The date in the LOGP filename is invalid.") from exc

    normalized_stem = re.sub(r"[_-]+", " ", stem).casefold()
    normalized_stem = re.sub(r"\s+", " ", normalized_stem).strip()
    if re.search(r"\bsold\s*out\b", normalized_stem):
        remarks = "Sold Out"
    elif re.search(r"\bsotex\b", normalized_stem):
        remarks = "Sotex"
    elif re.search(r"\bgood\b", normalized_stem):
        remarks = "Good"
    elif re.search(r"\bregular\b", normalized_stem):
        remarks = "Good"
    else:
        raise LogpConversionError(
            "The filename must contain Good, Regular, Sotex, or Sold Out for remarks."
        )

    return LogpFilenameMetadata(
        source_stem=stem,
        logp_no=logp_no,
        docu_date=docu_date,
        docu_name=docu_name,
        remarks=remarks,
    )


def _find_logp_header_row_and_columns(worksheet: Any) -> tuple[int, dict[str, int]]:
    required = {
        "item no": "item_no",
        "item description": "product_name",
        "uom name": "prod_uom",
        "warehouse asr": "warehouse_asr",
    }
    max_scan_rows = min(max(worksheet.max_row, 1), 15)
    max_scan_cols = min(max(worksheet.max_column, 1), 60)

    for row_index in range(1, max_scan_rows + 1):
        found: dict[str, int] = {}
        quantity_column: int | None = None
        for column_index in range(1, max_scan_cols + 1):
            normalized = _normalize_header(worksheet.cell(row_index, column_index).value)
            field_name = required.get(normalized)
            if field_name and field_name not in found:
                found[field_name] = column_index
            # The SAP file contains another Quantity column later. The approved
            # mapping uses the first Quantity column (source column F).
            if normalized == "quantity" and quantity_column is None:
                quantity_column = column_index
        if quantity_column is not None:
            found["record_qty"] = quantity_column
        if set(found) == {
            "item_no",
            "product_name",
            "prod_uom",
            "record_qty",
            "warehouse_asr",
        }:
            return row_index, found

    raise LogpConversionError(
        "Required LOGP columns were not found: Item No., Item Description, "
        "Quantity, UoM Name, and Warehouse/ASR."
    )


def _person_tokens(value: Any) -> list[str]:
    normalized = re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()
    return [token for token in normalized.split() if token]


def _resolve_logp_employee_name(
    source_name: str,
    employee_records: Iterable[dict[str, Any]],
) -> str:
    source_text = " ".join(str(source_name or "").split()).strip()
    if not source_text:
        raise LogpConversionError("Warehouse/ASR is blank in the uploaded LOGP file.")

    records = list(employee_records or [])
    if not records:
        raise LogpConversionError(
            "Master Data Employees is required to resolve Warehouse/ASR names."
        )

    if "," in source_text:
        surname_part, given_part = source_text.split(",", 1)
        surname_tokens = _person_tokens(surname_part)
        given_tokens = _person_tokens(given_part)
    else:
        source_tokens = _person_tokens(source_text)
        surname_tokens = source_tokens[-1:]
        given_tokens = source_tokens[:-1]

    all_source_tokens = set(surname_tokens + given_tokens)
    scored: list[tuple[int, str]] = []
    for record in records:
        official_name = " ".join(str(record.get("name") or "").split()).strip()
        if not official_name:
            continue
        variants = [official_name, *(record.get("aliases") or [])]
        best_score = -1
        for variant in variants:
            variant_tokens_list = _person_tokens(variant)
            variant_tokens = set(variant_tokens_list)
            if not variant_tokens:
                continue
            source_norm = " ".join(_person_tokens(source_text))
            variant_norm = " ".join(variant_tokens_list)
            if source_norm == variant_norm:
                best_score = max(best_score, 1000)
                continue
            if all_source_tokens and all_source_tokens.issubset(variant_tokens):
                score = 500 + len(all_source_tokens) * 10
                if surname_tokens and variant_tokens_list[-1:] == surname_tokens[-1:]:
                    score += 25
                best_score = max(best_score, score)
        if best_score >= 0:
            scored.append((best_score, official_name))

    if not scored:
        raise LogpConversionError(
            f'Warehouse/ASR "{source_text}" could not be matched to Master Data Employees.'
        )

    top_score = max(score for score, _ in scored)
    top_names = sorted(
        {name for score, name in scored if score == top_score}, key=str.casefold
    )
    if len(top_names) != 1:
        raise LogpConversionError(
            f'Warehouse/ASR "{source_text}" matched multiple Master Data employees: '
            + ", ".join(top_names)
        )
    return top_names[0]


def extract_logp_records(
    excel_bytes: bytes,
    employee_records: Iterable[dict[str, Any]],
) -> tuple[LogpSourceRecord, ...]:
    if not excel_bytes:
        raise LogpConversionError("The uploaded LOGP Excel file is empty.")

    try:
        workbook = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise LogpConversionError(
            "The uploaded file could not be opened as a valid .xlsx workbook."
        ) from exc

    worksheet = workbook.active
    header_row, columns = _find_logp_header_row_and_columns(worksheet)

    start_row: int | None = None
    fallback_row: int | None = None
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        item_no = str(
            worksheet.cell(row_index, columns["item_no"]).value or ""
        ).strip()
        product = str(
            worksheet.cell(row_index, columns["product_name"]).value or ""
        ).strip()
        if item_no.upper() == "A01AMB01":
            start_row = row_index
            break
        if fallback_row is None and item_no and product:
            marker = f"{item_no} {product}".casefold()
            if (
                "warehouse" not in marker
                and "grand total" not in marker
                and marker.strip() != "total"
            ):
                fallback_row = row_index
    if start_row is None:
        start_row = fallback_row
    if start_row is None:
        raise LogpConversionError("No valid LOGP product row was found below the headers.")

    employee_cache: dict[str, str] = {}
    records: list[LogpSourceRecord] = []
    for row_index in range(start_row, worksheet.max_row + 1):
        item_no_value = worksheet.cell(row_index, columns["item_no"]).value
        product_value = worksheet.cell(row_index, columns["product_name"]).value
        uom_value = worksheet.cell(row_index, columns["prod_uom"]).value
        quantity_value = worksheet.cell(row_index, columns["record_qty"]).value
        employee_value = worksheet.cell(row_index, columns["warehouse_asr"]).value

        if all(
            value in (None, "")
            for value in (
                item_no_value,
                product_value,
                uom_value,
                quantity_value,
                employee_value,
            )
        ):
            break

        item_no = str(item_no_value or "").strip()
        product_raw = str(product_value or "").strip()
        marker = f"{item_no} {product_raw}".casefold().strip()
        if "grand total" in marker or marker == "total" or item_no.casefold() == "total":
            break
        if not product_raw:
            raise LogpConversionError(
                f"Item Description is blank at source row {row_index}."
            )

        warehouse_asr = " ".join(str(employee_value or "").split()).strip()
        cache_key = warehouse_asr.casefold()
        if cache_key not in employee_cache:
            employee_cache[cache_key] = _resolve_logp_employee_name(
                warehouse_asr, employee_records
            )

        records.append(
            LogpSourceRecord(
                item_no=item_no,
                product_name=_remove_apostrophes(product_raw),
                prod_uom=None if uom_value in (None, "") else uom_value,
                record_qty=quantity_value,
                warehouse_asr=warehouse_asr,
                employee_name=employee_cache[cache_key],
            )
        )

    if not records:
        raise LogpConversionError("No LOGP product rows were captured.")
    return tuple(records)


def _load_logp_template(template_path: Path | None = None):
    path = Path(template_path or LOGP_TEMPLATE_PATH)
    try:
        if path.exists():
            workbook = load_workbook(path)
        else:
            template_bytes = base64.b64decode(LOGP_TEMPLATE_XLSX_BASE64)
            workbook = load_workbook(BytesIO(template_bytes))
    except Exception as exc:
        raise LogpConversionError("The LOGP conversion template could not be opened.") from exc

    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    headers = [worksheet.cell(1, column).value for column in range(1, 19)]
    if headers != LOGP_OUTPUT_HEADERS:
        raise LogpConversionError(
            "The LOGP conversion template headers do not match the approved format."
        )
    return workbook, worksheet


def _write_logp_output_rows(
    worksheet: Any,
    records: Iterable[LogpSourceRecord],
    metadata: LogpFilenameMetadata,
    process_date: date,
    auditor_name: str,
) -> int:
    records = tuple(records)
    style_prototypes = [
        copy(worksheet.cell(2, column)._style) for column in range(1, 19)
    ]
    alignment_prototypes = [
        copy(worksheet.cell(2, column).alignment) for column in range(1, 19)
    ]
    protection_prototypes = [
        copy(worksheet.cell(2, column).protection) for column in range(1, 19)
    ]

    clear_through = max(worksheet.max_row, len(records) + 1)
    for row_index in range(2, clear_through + 1):
        for column_index in range(1, 19):
            worksheet.cell(row_index, column_index).value = None

    login_date = datetime.combine(process_date, datetime.min.time())
    docu_date = datetime.combine(metadata.docu_date, datetime.min.time())
    logp_value: Any = (
        int(metadata.logp_no) if metadata.logp_no.isdigit() else metadata.logp_no
    )

    for row_index, record in enumerate(records, start=2):
        values = [
            None,
            login_date,
            logp_value,
            record.employee_name,
            docu_date,
            metadata.docu_name,
            None,
            None,
            None,
            None,
            record.product_name,
            record.prod_uom,
            None,
            None,
            record.record_qty,
            None,
            metadata.remarks,
            auditor_name,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column_index)
            cell._style = copy(style_prototypes[column_index - 1])
            cell.alignment = copy(alignment_prototypes[column_index - 1])
            cell.protection = copy(protection_prototypes[column_index - 1])
            cell.value = value
            cell.number_format = r"yyyy\-mm\-dd" if column_index in (2, 5) else "General"
    target_last_row = len(records) + 1
    if worksheet.max_row > target_last_row:
        worksheet.delete_rows(
            target_last_row + 1,
            worksheet.max_row - target_last_row,
        )
    return len(records)


def _safe_logp_output_filename(metadata: LogpFilenameMetadata) -> str:
    raw = f"FOR UPLOAD {metadata.logp_no} {metadata.remarks.upper()}.xlsx"
    return re.sub(r'[<>:"/\\|?*]+', "_", raw)


def build_logp_conversion(
    excel_bytes: bytes,
    filename: str,
    auditor_name: str,
    employee_records: Iterable[dict[str, Any]],
    *,
    process_date: date | None = None,
    template_path: Path | None = None,
) -> LogpConversionResult:
    conversion_date = process_date or philippine_today()
    clean_auditor_name = " ".join(str(auditor_name or "").split()).strip()
    if not clean_auditor_name:
        raise LogpConversionError(
            "The signed-in user's full name is required for auditor_name."
        )

    metadata = parse_logp_filename(filename)
    records = extract_logp_records(excel_bytes, employee_records)
    workbook, worksheet = _load_logp_template(template_path)
    _write_logp_output_rows(
        worksheet, records, metadata, conversion_date, clean_auditor_name
    )

    if workbook.calculation is not None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    buffer = BytesIO()
    workbook.save(buffer)
    output_bytes = buffer.getvalue()

    try:
        verification_book = load_workbook(
            BytesIO(output_bytes), data_only=False, read_only=True
        )
        verification_sheet = verification_book["Sheet1"]
        if verification_sheet.max_row != len(records) + 1:
            raise LogpConversionError(
                "Converted row count did not match the captured LOGP product count."
            )
        verified_headers = [
            verification_sheet.cell(1, column).value for column in range(1, 19)
        ]
        if verified_headers != LOGP_OUTPUT_HEADERS:
            raise LogpConversionError(
                "The converted LOGP headers failed the final integrity check."
            )
    except LogpConversionError:
        raise
    except Exception as exc:
        raise LogpConversionError(
            "The converted LOGP Excel file failed the final workbook integrity check."
        ) from exc

    return LogpConversionResult(
        output_bytes=output_bytes,
        output_filename=_safe_logp_output_filename(metadata),
        metadata=metadata,
        process_date=conversion_date,
        auditor_name=clean_auditor_name,
        records=records,
        source_signature=hashlib.sha256(excel_bytes).hexdigest(),
    )


def render_logp_conversion_page(
    user: dict[str, Any],
    employee_records: Iterable[dict[str, Any]],
) -> None:
    import pandas as pd
    import streamlit as st

    st.markdown(
        """
        <style>
        .iars-logp-hero {
            border: 1px solid #DDE5EF;
            border-radius: 16px;
            padding: 1.05rem 1.15rem;
            margin: 0 0 .9rem 0;
            background: linear-gradient(135deg, #F8FAFD 0%, #FFFFFF 58%, #FFF9EB 100%);
            box-shadow: 0 8px 24px rgba(6,26,54,.06);
        }
        .iars-logp-hero h2 { margin: 0; color: #061A36; font-size: 1.35rem; }
        .iars-logp-hero p { margin: .28rem 0 0; color: #667085; font-size: .88rem; }
        .iars-logp-route {
            display: grid;
            grid-template-columns: 1fr auto 1fr auto 1fr;
            align-items: center;
            gap: .55rem;
            margin-top: .85rem;
        }
        .iars-logp-route div {
            min-height: 64px;
            border: 1px solid #E4EAF2;
            border-radius: 12px;
            padding: .65rem .72rem;
            background: rgba(255,255,255,.9);
        }
        .iars-logp-route strong { display:block; color:#061A36; font-size:.86rem; }
        .iars-logp-route span { color:#667085; font-size:.74rem; line-height:1.25; }
        .iars-logp-route b { color:#C78B12; font-size:1.1rem; }
        @media (max-width: 760px) {
            .iars-logp-route { grid-template-columns: 1fr; }
            .iars-logp-route b { display:none; }
        }
        </style>
        <div class="iars-logp-hero">
          <h2>Sales Personnel LOGP Conversion</h2>
          <p>Convert an SAP LOGP export into the approved Sales Personnel upload template.</p>
          <div class="iars-logp-route">
            <div><strong>1. SAP LOGP Excel</strong><span>Upload the original LOGP stock file.</span></div>
            <b>→</b>
            <div><strong>2. IARS Mapping</strong><span>Resolve employee names, clean products and validate rows.</span></div>
            <b>→</b>
            <div><strong>3. Compatible Output</strong><span>Download the approved 18-column LOGP template.</span></div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        st.markdown("### Upload SAP LOGP File")
        st.caption(
            "Filename example: LOGP 61005123 DAVIDO GOOD - 06-22-2026.xlsx"
        )
        uploaded_file = st.file_uploader(
            "SAP Sales Personnel LOGP Excel",
            type=["xlsx"],
            key="sales_logp_excel_uploader_v4_5_13",
            help=(
                "The source file remains unchanged. Employee names are resolved "
                "from Master Data."
            ),
        )

    if uploaded_file is None:
        st.info(
            "Upload a LOGP .xlsx file. IARS will capture Item Description, UoM Name, "
            "Quantity and Warehouse/ASR, then map them to the approved template."
        )
        return

    auditor_name = str(user.get("full_name") or user.get("username") or "").strip()
    process_date = philippine_today()

    try:
        with st.spinner("Validating and converting the Sales Personnel LOGP file…"):
            result = build_logp_conversion(
                uploaded_file.getvalue(),
                uploaded_file.name,
                auditor_name,
                employee_records,
                process_date=process_date,
            )
    except LogpConversionError as exc:
        st.error(str(exc))
        return
    except Exception as exc:
        st.error(f"LOGP conversion failed: {exc}")
        return

    metadata = result.metadata
    employees = sorted(
        {record.employee_name for record in result.records}, key=str.casefold
    )
    metric_columns = st.columns(4)
    metric_columns[0].metric("Products Captured", f"{result.row_count:,}")
    metric_columns[1].metric("LOGP No.", metadata.logp_no)
    metric_columns[2].metric("Document Date", metadata.docu_date.isoformat())
    metric_columns[3].metric("Remarks", metadata.remarks)

    st.success(
        f"Conversion completed for {result.row_count:,} product rows. "
        "Product apostrophes were removed, employee names were resolved from Master Data, "
        "and the output passed the workbook integrity check."
    )

    with st.expander("Conversion Details", expanded=True):
        details = pd.DataFrame(
            [
                ["Login Date", result.process_date.isoformat()],
                ["Source Filename", uploaded_file.name],
                ["LOGP No.", metadata.logp_no],
                ["Document Date", metadata.docu_date.isoformat()],
                ["Document Name", metadata.docu_name],
                ["Remarks", metadata.remarks],
                ["Employee Name(s)", ", ".join(employees)],
                ["Auditor Name", result.auditor_name],
            ],
            columns=["Field", "Generated Value"],
        )
        st.dataframe(details, hide_index=True, width="stretch")

    st.markdown("### Converted Data Preview")
    preview = pd.DataFrame(result.preview_rows(limit=200))
    st.dataframe(preview, hide_index=True, width="stretch", height=390)
    if result.row_count > len(preview):
        st.caption(
            f"Showing the first {len(preview):,} of {result.row_count:,} converted rows."
        )

    st.download_button(
        "⬇️ Download Converted LOGP Excel",
        data=result.output_bytes,
        file_name=result.output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"logp_download_{result.source_signature[:16]}",
        type="primary",
        width="stretch",
    )
    st.caption(
        "Output format: Sheet1 · 18 approved columns · dates in yyyy-mm-dd · "
        "transaction, sales, invoice, product-code, invoice-quantity, "
        "discount-quantity and count fields remain blank."
    )


# ---------------------------------------------------------------------------
# Sales Personnel - Invoice Excel Conversion
# ---------------------------------------------------------------------------

INVOICE_OUTPUT_HEADERS = [
    "trans_id",
    "login_date",
    "logp_no",
    "employee_name",
    "docu_date",
    "docu_name",
    "sold_no",
    "inv_no",
    "pd_no",
    "prod_code",
    "prod_name",
    "prod_uom",
    "inv_qty",
    "disc_qty",
    "record_qty",
    "count_qty",
    "remarks",
    "auditor_name",
]

INVOICE_TEMPLATE_PATH = (
    Path(__file__).resolve().parent / "assets" / "invoice_conversion_template.xlsx"
)
# Embedded approved output-template fallback. This prevents deployment failures
# when a partial file replacement omits the assets directory.
INVOICE_TEMPLATE_XLSX_BASE64 = (
    'UEsDBBQABgAIAAAAIQBi7p1oXgEAAJAEAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslMtOwzAQRfdI/EPkLUrcskAINe2CxxIqUT7AxJPG'
    'qmNbnmlp/56J+xBCoRVqN7ESz9x7MvHNaLJubbaCiMa7UgyLgcjAVV4bNy/Fx+wlvxcZknJaWe+gFBtAMRlfX41mmwCYcbfDUjRE'
    '4UFKrBpoFRY+gOOd2sdWEd/GuQyqWqg5yNvB4E5W3hE4yqnTEOPRE9RqaSl7XvPjLUkEiyJ73BZ2XqVQIVhTKWJSuXL6l0u+cyi4'
    'M9VgYwLeMIaQvQ7dzt8Gu743Hk00GrKpivSqWsaQayu/fFx8er8ojov0UPq6NhVoXy1bnkCBIYLS2ABQa4u0Fq0ybs99xD8Vo0zL'
    '8MIg3fsl4RMcxN8bZLqej5BkThgibSzgpceeRE85NyqCfqfIybg4wE/tYxx8bqbRB+QERfj/FPYR6brzwEIQycAhJH2H7eDI6Tt7'
    '7NDlW4Pu8ZbpfzL+BgAA//8DAFBLAwQUAAYACAAAACEAtVUwI/QAAABMAgAACwAIAl9yZWxzLy5yZWxzIKIEAiigAAIAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKySTU/DMAyG70j8h8j31d2QEEJL'
    'd0FIuyFUfoBJ3A+1jaMkG92/JxwQVBqDA0d/vX78ytvdPI3qyCH24jSsixIUOyO2d62Gl/pxdQcqJnKWRnGs4cQRdtX11faZR0p5'
    'KHa9jyqruKihS8nfI0bT8USxEM8uVxoJE6UchhY9mYFaxk1Z3mL4rgHVQlPtrYawtzeg6pPPm3/XlqbpDT+IOUzs0pkVyHNiZ9mu'
    'fMhsIfX5GlVTaDlpsGKecjoieV9kbMDzRJu/E/18LU6cyFIiNBL4Ms9HxyWg9X9atDTxy515xDcJw6vI8MmCix+o3gEAAP//AwBQ'
    'SwMEFAAGAAgAAAAhAPHF0KLpAgAAeQYAAA8AAAB4bC93b3JrYm9vay54bWykVdtu4jAQfV9p/8Hye5oYQoCIUEGAXSS6i3p9QVqZ'
    'xBCLJM7aDlBV/fcdJ0AvvHRbBL6NdWbOzPHQu9xnKdoyqbjIA0wuHIxYHomY5+sA391OrA5GStM8pqnIWYAfmcKX/e/fejshN0sh'
    'NggAchXgROvCt20VJSyj6kIULAfLSsiMatjKta0KyWisEsZ0ltoNx/HsjPIc1wi+/AiGWK14xEYiKjOW6xpEspRqCF8lvFBHtCz6'
    'CFxG5aYsrEhkBUAsecr1YwWKURb503UuJF2mQHtPWmgv4evBj0CSKtY+HJ+5yXgkhRIrfQGwdh3wGXfi2IS8ob8/5/8xJNeWbMtN'
    '/U5RSe+TUXknLO8FjDhfRoOM9XsrnrL7WmmIFsUvmpnEphilVOlxzDWLA9yGrdixNweyLIYlT8HacJtNF9v9k/rmEsVsRctU34Lu'
    'jvAgZM/rNlrmJtRxkGomc6pZKHINsjkU76sSqbDDRIAg0TX7W3LJ4B2AJIArjDTy6VLNqU5QKdMAh/7iTgH9hdfsNjqLEVMbLYrF'
    'dHB9gyx0f0EWN4Or+WyMZr9/zBevFEbPpfwfGqORyYINaahDrdfvUwIRS/+oo7mWCNbT0QwKc0O3UCZQfHx4dFOoQ+fPUzgchoMW'
    'cazBZDK03NAbWYOGN7CGhIBp2B11WuEzsJCeHwla6uRQeoMZYBfqfGa6ovujhTh+yeMX/0/O4WOZ+d1wtD0bpqYn3XO2Uy8iMVu0'
    'f+B5LHYBtkgD2Dy+3e4q4wOPdQIqc9pNuFKf/WR8nUDEhDiuU6XylYeqmYGnakZ5pehpvq0kAa0PlUUqaIxGdMtjAY3U9D6TP4KR'
    '9Dks5DQmFegRKaJpBKI2k7lYeWR7PVO634MZpMQD/DQibtdpjgdWsxm6ltuetK3OxGlZTbfthi13OCZO2yTftGZ/n+6i7eeecMO1'
    'j30+fN0jD7U0Qjfg/uEPBCmmDybD0egOYq7HisEJrf8PAAD//wMAUEsDBBQABgAIAAAAIQCBPpSX8wAAALoCAAAaAAgBeGwvX3Jl'
    'bHMvd29ya2Jvb2sueG1sLnJlbHMgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsUk1LxDAQvQv+hzB3m3YV'
    'Edl0LyLsVesPCMm0KdsmITN+9N8bKrpdWNZLLwNvhnnvzcd29zUO4gMT9cErqIoSBHoTbO87BW/N880DCGLtrR6CRwUTEuzq66vt'
    'Cw6acxO5PpLILJ4UOOb4KCUZh6OmIkT0udKGNGrOMHUyanPQHcpNWd7LtOSA+oRT7K2CtLe3IJopZuX/uUPb9gafgnkf0fMZCUk8'
    'DXkA0ejUISv4wUX2CPK8/GZNec5rwaP6DOUcq0seqjU9fIZ0IIfIRx9/KZJz5aKZu1Xv4XRC+8opv9vyLMv072bkycfV3wAAAP//'
    'AwBQSwMEFAAGAAgAAAAhAAMTAIjuEQAAm4QAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWyUkk1v2zAMhu8D9h8E3WvZbtIk'
    'RpyiWBCst6Hr1rMs07EQfXgS0yT/frSSDAV6yS4mZVPPS7708vFoDXuHELV3NS+ynDNwyrfabWv+63VzN+csonStNN5BzU8Q+ePq'
    '65flwYdd7AGQEcHFmveIQyVEVD1YGTM/gKMvnQ9WIh3DVsQhgGzTJWtEmecPwkrt+JlQhVsYvuu0grVXewsOz5AARiL1H3s9xCvN'
    'qltwVobdfrhT3g6EaLTReEpQzqyqnrfOB9kYmvtYTKS6stPhE95qFXz0HWaEE+dGP8+8EAtBpNWy1TTBaDsL0NX8qahe5g9crJbJ'
    'oN8aDvFDzlA2P8GAQmhpT5yN/jfe78bCZ3qVEzKmghEpFep3+AbG1Hxd0gr/JBFKSUD8U/iYX9U2aWM/Amuhk3uDL/7wHfS2R5Kd'
    'kgOjEVV7WkNUtAESzsrpSFXeEIKezGr6lUjUymOKB91iT7fzbFbki/sZURqIuNEjkjO1j+jt26XogjpDJhcIxQukvM8m5XQ2L0j0'
    'VgpVplYo/ncrIo31FwAA//8AAAD//5Sd3ZIbtxWEX8Wle0o7w+HPumRVRZEdJ+Q6od9ApajKV07KUinJ2wcAsTvT3UADuEtyTh/s'
    'Nk8w+OZgqbdffvv8+euHj18/vnv7x7/+890fP7yaXn335d8ff/8S/tP30/nVd/+dlo+fvv/n/z58/vLp8+9ff3j18Ho+vHr39lNM'
    '/lPIDv/Tl/Dfv717ePvm27u3bz7l2PttbMLYn7exGWMftrE9xn7cxhaM/bSNHTD2l23siLGft7ETxv66jZ0x9rdt7BFjF/jdyZgr'
    'BMmZJwiSNb9AkLz5OwTJnH9AkNy5QZDs+RWCqz9vQr+8NM080jTvY3ZstNg0y3F+4OaYU+Q4PTwcppl+yw9B/NJxE30kP2Ll6ZF+'
    '0Z9ATJ/Zz/dlH8+Phz19YpetbuaPE4L8cd6L7ugHvYGGPuVfIbj+/uD4fsjxmO0c3zvHg7juOFZWx0HMjt+XjY7PZNBlqxPHIciO'
    '34vuaK0baNhxCFYcX4Ycj9nO8cU5HsR1x7GyOg5idvy+bHKcPLhsdeI4BNnxe9HdRDvHDURsOQQrlh+GLI/ZzvL41Pr2rrKtBHHd'
    'cqysloOYLb8vmywngy5bnVgOQbb8XnS3p93xBiK2HIIVy49DlsdsZ/nRWR7EdcuxsloOYrb8vmzaycmgy1YnlkOQLb8X3S1n2v9v'
    'oGLPIVjx/DTkecx2np+c50Fc9xwrq+cgZs/vyxY93+rEcwiy5/eiu8PCz09QsecQrHgeTrb9x9z3Mdt5fnaeB3Hdc6ysnoOYPb8v'
    'm7YWOulctjrxHILs+b3o7pF3cxCx5RCsWP44ZHnMdpY/OsuDuG45VlbLQcyW35ctWr7VieUQZMvvRXfTA2/noGLPIVjxfHoYMj2l'
    'O9dDgnmIRnnddyquxqOcnc9LJ+sFt7YLi/dQdxbguv9Ku4lPjChj9zFasz/A08A2M8V0a/8dmipnmCg39mPxgv0gF/vvSyf7+aEK'
    'C6v927pq/73ujvy9YU1xH2rW3B/D0qnFpSHBNb8lUypecN+yaV46PV4J7S+x9MvHru5bPM11d0d5JWAJFZes2T/GqFMLUkOCs99i'
    'KhUv2G9BNS+d7GduiqWN/ZZVc93dtCx8qsSy0v89xDqNIWtKt7uPhdYoN7tPC1tRLrvPCq57+QAsuUJd3X2e2ZWeCjeUif099DqN'
    '4WtKt/ZbgI1yY38LYVEu9m8gll6RXkCp24/F2KCN/5feTbL7W5DFJWvbzxjKTi2WDQlu+7E0S8UL24/l2bx0evZSo15iabP9WKTN'
    'dfXZa5EWV6y5v0Lt3J4jvJ9aVBsSnPuWa6l4wX1Ltnnp5L48ey3bxoXXz0YOnplu9Y0C6mTz6eHbOLrJJ88u/1uEG+o5/y3jpp9l'
    '3dkK/lvKzUuX/becGxc2/mfSpc/0hipxvwd1p5V1u9xvwW6o59y3uJt+Fuu+Bd68dDr6yN5jkTcubNzP0PsoBx8LvVi0svnMK/X2'
    '2J/S7UDKUm+U15+8VLwwkwK5DKXuSxdnJLDwTG18xSid759CND15wzQXR5c31HH7Y7Tm/4q9Xf63sHe22Buizv8W9qJc/N9gL7++'
    'B6X6D4gq/lewF2uK+z3YO6/Y2+V+cxzr57F+INucyPqR7DqTnfkNZ/w11+1Fuh+i4n4ey0rv+7lsz2B2Xqm3y/0W9YZ6ZuuPq5m9'
    'pzWdRbn0/mY+y698QKm9D3gq7ucR7fSaH71YVbq/B3rnFXq7/G8NakM957+F3vSzuEdvTFg/PvHfQC8o1X/gU/E/Q+/C0ItFxf4e'
    '6A13g0bOnSndPnot9Ea5af8W9KJc7DfQC0q1H+hV7M/Q+/CaX7phVfG/Z347r9Tb1f4t6g31XPtb6k0/i21/S7156SL1xtJm84eo'
    '+J8nudL9lnpxxdrBZ4x65xb1hgTnvqVeKl44eFrqzUuX3QcClUcvRMX9TL3ivh3pxl9m/bxr7o8x79xi3pDg3LfMS8UL7lvmzUsX'
    'mTeWNr0PUXE/M+9JH712votr1vwfo965Rb0hwflvqZeKF/y31JuXLlJvLG38h6j4n6l31r3fYi+uWfF/P4a9Kd09e0OC8T/K689e'
    'Kq7+o5yfvXnp4nUSUM4EZVeMyuXAjL1yuwFl/OjFaM3+Merdt6g3JDj7LfVS8YL9dtibly5u/rH02v5ivx325rrhEhW9dMCiYn8P'
    '9u7HsDel2+632Bvlpvtb2Ity6f7NVWR+5wZK7X477Q3a9NJH7bfci0vWun+Me/ct7g0Jrvst91LxQvfbaW9eunjJJ5Zeu59mYVeM'
    '8uaf68q4BVXS/D3Uux+j3pRum99Sb5Sb5m+NelEuzW/uKIOS74xfMUp1n0I0NT/fa0OVuN8Dvfsx6E3p1n0LvVFu3G9BL8rF/RV6'
    '+XLxBZTqPuCpuJ+h98Cv+7Go2N/DvPsx5k3p1n7LvFFu7EegLmw9lnlD8dikxVtWsLDaD/Qq9mfmlea3zIsr1jb+Mebdt5g3JLiN'
    '3zIvFS+4b5k3L1285hNLrxs/nV+uGBX3M/Py7YgbyqT5e6B3Pwa9Kd02v4XeKDfNj0RdsN9Cbyj+0vx8ywQW1uYHeBX7M/Ty236s'
    'Ke73DHr3Y8ib0q37Fnmj3LiPPF1w3yJvKP7iPl9zgIXVfUBXcT8jr2w9FnhxxcrWs4wBb0p37ocEs/VEed19Kq7uo5yfu3np4gsH'
    'UIr7GGX3c91wt5yIC2Xc/Bit2T8GvEsLeEOCs98CLxUv2G+BNy9d3Plj6ZfPnf809ArRWezPY161H5BW7O8B3mUMeFO67X4LvFFu'
    'ur8FvCiX7t/MeSl2AeWecPgKUbU/Ay+/bEaVuN8z513GeDelW/ct70a5cb8150W5uL+Z8/LrBlCq+8Cm0vzPc17eekAl7vfw7jLG'
    'uynduu//HtfyLhUvbD12yhvkz89d9vcSS69bj/Q+sKm4/3y1md0Hlbjfw7vLGO+mdOu+5d0oN73f4l2US+8b3gUlXzu/QlQulodo'
    'etsgL5pRJvb38O4yxrsp3dpveTfKjf0t3kW52L/5U13y4gJKtR/Ild/zB21+0ynnHku8sObmuyfguxeWMeJN6dZ/S7xRbvxHnC5s'
    'PpZ4Q/GXzYe/9AIW5pdBV4jy5aynXHcnD1475cWatVPnGPAurSlvSHCnTgu8VLzgvgXevHTxcmcsvW79dP3qClF1PwOvXu5Enew+'
    'Pci7jCFvSrfdb5E3yk33t5AX5bL7mL/mBaV2v53yBm3afehTu0HNWdyHmpXuP4whb0p37ocE0/1RXnefimv3o5zdz0sXmQuU4j5E'
    'pftz3d3E16tQxvZjtGb/GPIeWsgbEpz9FnmpeMF+i7x56eIVk1i6vvlAVO3PyCvuW+LFmjX3x4j3gFCqXzQVEpz7lnipeMF9kEvz'
    'mxFvLG3cBzrlGWP+lXbivp3wwoq1Y89hjHhTut16LPFGudl6WsSLcnF/JV5hLlDu+XIbROXUH6LlUz/KZOvpQd7DGPKmdGu/Rd4o'
    'N/a3RrwoF/s3I16+WA7KPbXxFaP0SvkpRJP9/KIfVPLcxWht6xlD3gNSaWHr8d9GZZGXihe2HpCL+5t7zcxcsfS69fD1Bohq82fk'
    '5evQN5RJ8/cg72EMeVO6bX6LvFFumr+FvCgX+zcjXrF/u/Be7LfIG1a9Iy+974GfRpsfataafwx4D60Rb0hwz10LvFS80PwWePPS'
    'xQF7LG2aH9CVXzjkuvGbHNh/i7ywZvXJO4a8hxbyhgTnv0VeKl7w3yJvXrrsPyCvdD/Aqfj/POMV+0Emm08P8R7GiDel283HEm+U'
    'm82nRbwol81nJV4+2lxAqZsP0KnY/3yvWb7HBKrq9tPDvMcx5k3pzv+QYNo/yuv+U3Ftf5Sz/3npIvOCUvyHqDx7c93dLP6jjvsf'
    'o5Xt/zgGvSnd+m+hN8qN/0jUBf8t9Ibi1WELLKz+A75y/+e6he0fqkr/Y7Tm/xj2HlvYGxJc/1vspeIF/y325qWL1xxi6frjF6La'
    '/89/0MvbP8qk/Xsmvccx7k3ptv0t90a5af8W96Jctp8N9zJ4gXLPF9wwyuAVounsyW8dQKXN34O9xzHsTenWfYu9UW7cb2EvysV9'
    '8/e8oFT3YQ4s7j9PeuWbae2oF5asHT2PY9yb0q39lnuj3NjfGvWiXOzfjHoZvEC5J+UVovLCM0RT88sdH5TJ3tPDvccx7k3p1n7/'
    'jcyWe6l4Yeu3V5uD/PnRy28ILrH0uvWL/cCo/MYz193xsAVq6t7Tw73HMe5N6dZ9y71Rbpq/NehFuTT/ZtDLd3xAyQfIK0T3svc8'
    '/zkvv3MDmdrfc7X5OIa9Kd3ab7E3yo39yNSF5rfYG4rXz51bJX8nwxV+LLW/hr0gU/t7sPc4hr0p3dpvsTfKjf0t7EW5dP9m0Mv/'
    '0gQoF2rjK0R168/Y+/Badh97uxmrVo79pzHsTenO/5Bgjv1RXvefimv7o5z9z0sX3/qAUvyHqPif68pf1KGKn7wYrbk/Br2n1qQ3'
    'JDj3LfRS8YL7Fnrz0sWXDrH0y8eu7gP08pM319W75VBUNh+M1uwfY95Ti3lDgrPfMi8VL9hvmTcvXRy0x9LGfjvqzXVl1As11f0e'
    '5D2NIW9Kt1uPRd4oN1tPC3lRLltPRt5lWk586gflwnesICpvHEL0Pm6Rf5rC3m7GorXmH2PeE2KpDhtDgmt+y7xUvND89nZzXvox'
    '2s9/VRRLr80v9gO88gu3XHd3kr+pg6ra/j33m09j0JvSbftb6I1y0/4t6EW5tH+G3ug/v/EBpbY/8Kn4/zLs5ZcOUFX976He0xj1'
    'pnTrv6XeKDf+t6a9KBf/M/VG/5m7QKn+22lv0Fa2H5DJ0acHe09j2JvSrf3+nySy2EvFC9uPHfcGeeKuaD/fb46lzfZjx725rn6R'
    'CRTV7u/B3tMY9qZ0a7/F3ig33d/CXpRL92fsDfYfKXYBpXa/nfYGbaX77bQXlqy98TyNYW9Kt/Zb7I1yY38Le1Eu9mfsLT58twvz'
    'pYUr1NWzT8Ze+eMWlMnm0zPsPY9Rb0p39ocEc/aJ8rr9VFw3H5Sz/Xnp4tkHlAtfc4Oo2J/r7uYznz1Rx/5jtHL2PI9xb0q3/lvu'
    'jXLjf2vYi3LxPw97S2cfUKr/dtgbtGn3Kfhv7zjDmrXt5zwGvind+m/BN8qN/0jVhf634BuKPz98j/x9DrCw+g+QymfPXDec/emq'
    'FRSVhy9Ga+0/Rr5nhFNFr5Dgth9LvlS8YD/Ipf035MvoFUuvZx/+EjeM0kf3lH8lee8AKnW/Z9h7HgPflG6b34JvlJvmbw17US7u'
    '52Fv3Hxoe7+AchH3AYvF/TzslS9SgqJqfw/3nse4N6Vb+y33Rrmxv8W9KBf7V+498ht/UPJ34V0hqs/e2t/1okwevT3Yex7D3pRu'
    '7bfYG+XG/hb2olzsX7FX7d8uvPCwF+rKFzqEaHr08rgFVeK+pd43X377/Pnrh49fP777PwAAAP//AAAA//+yKUhMT/VNLErPzCtW'
    'yElNK7FVMtAzV1IoykzPgLFL8gvAoqZKCkn5JSX5uTBeRmpiSmoRiGespJCWn18C4+jb2eiX5xdlF2ekppbYAQAAAP//AwBQSwME'
    'FAAGAAgAAAAhAMEXEL5OBwAAxiAAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7FnNixs3FL8X+j8Mc3f8NeOPJd7gz2yT3SRknZQc'
    'tbbsUVYzMpK8GxMCJTn1UiikpZdCbz2U0kADDb30jwkktOkf0SfN2COt5SSbbEpadg2LR/69p6f3nn5683Tx0r2YekeYC8KSll++'
    'UPI9nIzYmCTTln9rOCg0fE9IlIwRZQlu+Qss/Evbn35yEW3JCMfYA/lEbKGWH0k52yoWxQiGkbjAZjiB3yaMx0jCI58Wxxwdg96Y'
    'FiulUq0YI5L4XoJiUHt9MiEj7A2VSn97qbxP4TGRQg2MKN9XqrElobHjw7JCiIXoUu4dIdryYZ4xOx7ie9L3KBISfmj5Jf3nF7cv'
    'FtFWJkTlBllDbqD/MrlMYHxY0XPy6cFq0iAIg1p7pV8DqFzH9ev9Wr+20qcBaDSClaa22DrrlW6QYQ1Q+tWhu1fvVcsW3tBfXbO5'
    'HaqPhdegVH+whh8MuuBFC69BKT5cw4edZqdn69egFF9bw9dL7V5Qt/RrUERJcriGLoW1ane52hVkwuiOE94Mg0G9kinPUZANq+xS'
    'U0xYIjflWozuMj4AgAJSJEniycUMT9AIsriLKDngxNsl0wgSb4YSJmC4VCkNSlX4rz6B/qYjirYwMqSVXWCJWBtS9nhixMlMtvwr'
    'oNU3IC+ePXv+8Onzh789f/To+cNfsrm1KktuByVTU+7Vj1///f0X3l+//vDq8Tfp1CfxwsS//PnLl7//8Tr1sOLcFS++ffLy6ZMX'
    '333150+PHdrbHB2Y8CGJsfCu4WPvJothgQ778QE/ncQwQsSSQBHodqjuy8gCXlsg6sJ1sO3C2xxYxgW8PL9r2bof8bkkjpmvRrEF'
    '3GOMdhh3OuCqmsvw8HCeTN2T87mJu4nQkWvuLkqsAPfnM6BX4lLZjbBl5g2KEommOMHSU7+xQ4wdq7tDiOXXPTLiTLCJ9O4Qr4OI'
    '0yVDcmAlUi60Q2KIy8JlIITa8s3eba/DqGvVPXxkI2FbIOowfoip5cbLaC5R7FI5RDE1Hb6LZOQycn/BRyauLyREeoop8/pjLIRL'
    '5jqH9RpBvwoM4w77Hl3ENpJLcujSuYsYM5E9dtiNUDxz2kySyMR+Jg4hRZF3g0kXfI/ZO0Q9QxxQsjHctwm2wv1mIrgF5GqalCeI'
    '+mXOHbG8jJm9Hxd0grCLZdo8tti1zYkzOzrzqZXauxhTdIzGGHu3PnNY0GEzy+e50VciYJUd7EqsK8jOVfWcYAFlkqpr1ilylwgr'
    'ZffxlG2wZ29xgngWKIkR36T5GkTdSl045ZxUep2ODk3gNQLlH+SL0ynXBegwkru/SeuNCFlnl3oW7nxdcCt+b7PHYF/ePe2+BBl8'
    'ahkg9rf2zRBRa4I8YYYICgwX3YKIFf5cRJ2rWmzulJvYmzYPAxRGVr0Tk+SNxc+Jsif8d8oedwFzBgWPW/H7lDqbKGXnRIGzCfcf'
    'LGt6aJ7cwHCSrHPWeVVzXtX4//uqZtNePq9lzmuZ81rG9fb1QWqZvHyByibv8uieT7yx5TMhlO7LBcW7Qnd9BLzRjAcwqNtRuie5'
    'agHOIviaNZgs3JQjLeNxJj8nMtqP0AxaQ2XdwJyKTPVUeDMmoGOkh3UrFZ/QrftO83iPjdNOZ7msupqpCwWS+XgpXI1Dl0qm6Fo9'
    '796t1Ot+6FR3WZcGKNnTGGFMZhtRdRhRXw5CFF5nhF7ZmVjRdFjRUOqXoVpGceUKMG0VFXjl9uBFveWHQdpBhmYclOdjFae0mbyM'
    'rgrOmUZ6kzOpmQFQYi8zII90U9m6cXlqdWmqvUWkLSOMdLONMNIwghfhLDvNlvtZxrqZh9QyT7liuRtyM+qNDxFrRSInuIEmJlPQ'
    'xDtu+bVqCLcqIzRr+RPoGMPXeAa5I9RbF6JTuHYZSZ5u+HdhlhkXsodElDpck07KBjGRmHuUxC1fLX+VDTTRHKJtK1eAED5a45pA'
    'Kx+bcRB0O8h4MsEjaYbdGFGeTh+B4VOucP6qxd8drCTZHMK9H42PvQM65zcRpFhYLysHjomAi4Ny6s0xgZuwFZHl+XfiYMpo17yK'
    '0jmUjiM6i1B2ophknsI1ia7M0U8rHxhP2ZrBoesuPJiqA/a9T903H9XKcwZp5memxSrq1HST6Yc75A2r8kPUsiqlbv1OLXKuay65'
    'DhLVeUq84dR9iwPBMC2fzDJNWbxOw4qzs1HbtDMsCAxP1Db4bXVGOD3xric/yJ3MWnVALOtKnfj6yty81WYHd4E8enB/OKdS6FBC'
    'b5cjKPrSG8iUNmCL3JNZjQjfvDknLf9+KWwH3UrYLZQaYb8QVINSoRG2q4V2GFbL/bBc6nUqD+BgkVFcDtPr+gFcYdBFdmmvx9cu'
    '7uPlLc2FEYuLTF/MF7Xh+uK+XNl8ce8RIJ37tcqgWW12aoVmtT0oBL1Oo9Ds1jqFXq1b7w163bDRHDzwvSMNDtrVblDrNwq1crdb'
    'CGolZX6jWagHlUo7qLcb/aD9ICtjYOUpfWS+APdqu7b/AQAA//8DAFBLAwQUAAYACAAAACEAYpngc1wIAABdQQAADQAAAHhsL3N0'
    'eWxlcy54bWzEXFtv4kYUfq/U/2BZ6iPxHXAErAKJ25W26Uqbqn01xhA3viB72JJW/e89MzaeMwHCkDiMot3FDvOd63xn5tizo0/b'
    'LNW+x2WVFPlYt65MXYvzqFgk+Wqs//4Q9Ia6VpEwX4Rpkcdj/Tmu9E+TH38YVeQ5jb89xjHRACKvxvojIetrw6iixzgLq6tiHefw'
    'm2VRZiGBy3JlVOsyDhcVHZSlhm2afSMLk1yvEa6zSAYkC8unzboXFdk6JMk8SRPyzLB0LYuuP6/yogznKai6tdww0rZWv7R3Etit'
    'PSFZEpVFVSzJFYAaxXKZRPG+rr7hG2HEkQD2bUiWZ5h2bfhktCxyUmlRsckJuB+czVS8fsqLv/OA/g7u6vXXJqPqH+17mMIdSzcm'
    'o6hIi1Ij4G0wlt3JwyyuvzEL02ReJvRryzBL0uf6tk1vsAA138sScBe9aVBFlMkZ7tnj0Dt79mhfktUjOW1V+NcBq+bU9p0HPUmJ'
    'b/SgIIvZIkTrsHVdyNrPjE5lHcrAcjUf60EAs9kyzYNh6zANG2H+zAR5FxPmDS5mmRM4waBTy4RcRPnRuJIKdIIuXXlCYHAzuL2Y'
    'O7sXdsy6hogvNQPohOvWiwnmx/08GQT05xJ5cuGS1lHEWAWtoIQmadpWdMehxRvuTEawWiFxmQdwoTWfH57XULpzWFhRtxr19058'
    'e1WGz5bNypfcgKpIkwXVYjVjC4Zm2s/6d8HsjslFmslqcQQ0CGaDDwC9m/qz7jWd+X7XoHYAPx2D3nj0p3PzIVSd+bSZsG5XSrZ4'
    'Gknouti8Gvi+P7T6w+HQdx3LdZmT501GJ/ki3saLsd7vzE37Gnigge8M/b4NipjukIm6qAYOKDDwvKFn+bYLfxhFf7wGXfvU01VH'
    'FWmgKKpIA0VRZStMowPmb2ZKX3lUkQaKooo0UBTVQccMPFAeVaSBoqgiDRRFlXVHOpyr0GVSXFeRBoqiijRQFNXOFp8NA/vKo4o0'
    'UBRVpMG7o8p2V7CfmxflAvrjbY/WhL1UfW8ySuMlgZ1bSbuS8C8p1vD3vCAE+siT0SIJV0UepnRvtxshMRL67dBaH+vkMYmeQJjQ'
    'OazX2LWIj5LQsoNLV93uwDUHrmf3641NR6KzeJFssn3rWtkH4wdupL49bTjyYd4KaXa+vJFh0Pg14ZMcwULNIi05AHJilxKSI7qw'
    'kTf1ZG1EI+RsRAMkbUQjOsqiRbGBRz0vAxwEQ9Nku6Oz8+Uw4OvePDlm358nhxzw6MkxXeTN1KY/bBUpOTfQCLm8QQMk8waNeFve'
    'HJx4QtPitL+Fr7+mRkP1UDmiOE2/US7/c9mWD+hcTEbbpZZvsiAjn6FpAU9c6bO33UdoDzYf61JRX0AJOTYIHm7SluGBQVq4XqfP'
    '95tsHpcBewzLpLG7tA3Jr6asxvHrmzRZ5VnMei96DfO1LEgcEfaYmHVMj+njHNHHaoBk9HmPfPeIfPCTtD/eIx/29gfjAX5RKh/y'
    'TEk+wK545w8IAU7q1/TpMiNhB7fTAIKgQgPYbew0gPRUoQGsjHcaQIJyDUCdV7LiPfPAosTWEBPkABcJ8j9KJHDMQZEfaOUx+gWT'
    'LzLdLcS3kOjczXDxUW4+RrGqKAaFHWYadwFcvBYDqLrdlDzrGOcrcwgiXdCBewRo4DJZiTgXZKpgPJQUlIpUqwDqqFZBUfm1UC5Q'
    'tlLtBkUVGOejUIJfJ4kuF0JYBaEkX1AFnAxCvVLjBqFeqFFBPT/aqvgRJQM8zlbNDLYqgkQ7JPX8aKviR5wM6gnSVkWQKBnU86Ot'
    'ih9xMqgnSEf9AtJRz4/OB/OjgRumdfsUdU7f1jjVtsvDHVSLbmJP9l3b4Ye2TfUjBqo0qInavWKztzVKo8cZxrpt/qT1tJsogvYq'
    'xLTWz6I5vklSeC5P1aX5Fm0qeNQ1rW82Bzxew6KHXJitNq0hCAsMPRcLEBosSkEIC+jgXCwQX2OxDOZY0Ak/GwvKQoNFCwTCArI+'
    'Vy8Y0mCJvvckfe8eiiNbSSF/QYBl9MJYPI60ECMsMPlcLB5HyuMIC0w+F4vHkdIA8j0IOReLx5FWWYQF6XYuVhtHl5I0x/Ikfd8/'
    'GEfKtshfkrmKsXgcxVx1JHMVY/E4irlKTZbxF8bicRR5wpXkCYzF4yjyhCvJExiLx1H0vSfp+5eMKma8LZnxNQqPnZjr8Dq/lL9r'
    'FB41McsdySyvUXi8xPx2JfO7RuGRElnFlWSVGoVzpuhdV9K703CxY10xYWxJl8C5xmiTwrnPgp4aZeUODkPiqUo3cTJTYvYYR0/a'
    'DIp3CyTOB1pGZYDutus0zENSlM/aQ7wlLZwYdE8S7ueiaH0kItAdiYxCv8AZWzi+q7WLCzGHrTNh2rkguod29c7Rpp0MYv7RI6/n'
    'wLSzQSRVun6SgfmcrzdthEQupaVbBuJLkj/FCzFzRA/THbQM0n28IWXY5t8LwpJ0zD19sN9iiBTBntO/XDPew3P8NkdfeFHSBb9t'
    'CHIj667zWkkbOTLWPyQEXt3ZTWJhDtPnKVIQBeGm04REJRuOSElh/BGWOZ0twtR9kaNHLOL7FVj9L7b8JQ/md0JPobPXP9r9ALh3'
    'ES/DTUoe2l+Odf75V/ZSHCRT862vyfeCMIixzj+zM9Awi+GdEKCbLxW8wQb/apsyGev/3k0H/u1dYPeG5nTYc53Y6/ne9LbnubPp'
    '7W3gm7Y5+w8dh3/HYXh2dB/eS7Hc6yqFI/NlY2yj/Dd+b6yji1p99ooRqI119+2+eeNZZi9wTKvn9sNhb9h3vF7gWfZt353eeYGH'
    'dPfeePzeNCxrd/x+a3nXJMniNMl3sdpFCN+FIMHlK0YYu0gY/P9FmPwPAAD//wMAUEsDBBQABgAIAAAAIQBFPCgwKAIAAAUGAAAU'
    'AAAAeGwvc2hhcmVkU3RyaW5ncy54bWxsVF1v2zAMfB+w/0AY2NOQ2OmaoRiSFMm67itd27TLQ18CxeJcrZLoSnTg/Psp6bABNh99'
    'R/LuaEKT89ZZ2GGIhvw0Gw2LDNCXpI2vptnP+8vBWQaRldfKksdptseYnc9ev5rEyJB6fZxmj8z1hzyP5SM6FYdUo0/MLwpOcfoM'
    'VR7rgErHR0R2Nj8pive5U8ZnUFLjeZqNT5JM481zgx//IkU2m0Qzm/CMg/JxY/Qk59kkP2AvuKXK+I1WjAJTbzx1YXS1pT3ixivX'
    '69FUNuKwIyF1RLJaUDF+J6C1VFoH0pu07J6ZIyFpHomGXDfaQfSZ911Ym1hKeMCSgpaY4w+RW5wKT7GroBptmIK40hXuYa08LFUT'
    '0DPChdoZ3fstX/2OTNnbwdxtA7WGjYUr1cK4cj3xqNLJcdfSZ6LerXxDnWxEDF51y+eO9y8So6LoSWyp7TYsB3c1lqKhH5eDa48g'
    'DbptTPkEX1BZGBfOdoduidn2dnAQ4kAu5R/AEuBtGt3vlaqEsnusqYXTM9H5PLKpROPXwZv0ArDYdt3uYWU0jPuudkb1Uj5gILgy'
    '6RSkHAvD6hBV4m7SFigdGpy+E12uMNamQg+jMehAdZQsrdCS0i/n1De8Jsu1xRa+34oSCxU5vYEpwY1t0nzhIBcBUWOA+cUnkU8p'
    'ksfBA5wUQvc/dvRUde/jtwpdaG226e0rOUV+A3dkh+LihKrl/0l5esZnfwAAAP//AwBQSwMEFAAGAAgAAAAhAEKZs81AAQAAbQIA'
    'ABEACAFkb2NQcm9wcy9jb3JlLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJySX0vDMBTF3wW/Q8l7'
    'm3TTMULbgcoexIGwieJbSO66YPOHJNrt25u2W53MJx+Tc+4v51xSLPaqSb7AeWl0ifKMoAQ0N0LqukQvm2U6R4kPTAvWGA0lOoBH'
    'i+r6quCWcuPg2RkLLkjwSSRpT7kt0S4ESzH2fAeK+Sw6dBS3xikW4tHV2DL+wWrAE0JmWEFgggWGO2BqRyI6IgUfkfbTNT1AcAwN'
    'KNDB4zzL8Y83gFP+z4FeOXMqGQ42djrGPWcLPoije+/laGzbNmunfYyYP8dvq6d1XzWVutsVB1QVglPugAXjqkejoEnWTAfJalPg'
    'M6lbY8N8WMWNbyWIu8OF+9IR2X2V4QEQSQxHhyon5XV6/7BZompCJrOUzFNysyGE3hI6zd+7AL/mu7DDhTrG+DfxBKgKfPFBqm8A'
    'AAD//wMAUEsDBBQABgAIAAAAIQAJN7BhjgEAABEDAAAQAAgBZG9jUHJvcHMvYXBwLnhtbCCiBAEooAABAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'
    'AAAAAAAAAAAAAAAAAAAAAJySQWsbMRCF74X8B6F7rHUSQjFahRI3pNBSg530rEqzXhFZEtJ4sfvrO7tLnHXSU2+jmcfTpzeSd4ed'
    'Zx3k4mKo+XxWcQbBROvCtuZPm4fLz5wV1MFqHwPU/AiF36mLT3KVY4KMDgoji1Bq3iKmhRDFtLDTZUbjQJMm5p1GOuatiE3jDCyj'
    '2e8goLiqqlsBB4RgwV6mkyEfHRcd/q+pjabnK8+bYyJgJb+k5J3RSK9UP5zJscQG2deDAS/FdCiJbg1mnx0eVSXF9CjXRnu4J2PV'
    'aF9AireGfATdh7bSLhclO1x0YDBmVtwfiu2Ks9+6QI9T805npwMSVi8bD0PtU8GsfsX8UloALFKQYGwO5VQ7rd2Nmg8CKs6FvcEI'
    'QoNzxI1DD+Vns9IZ/0E8nxIPDCPviPMtdJFWyWi7bJ981JYtdeds/IA8pECXv7vuuwsv5Slt4lIjvMZ53pTrVmewtIFT3KeGfKQk'
    's+9N7lsdtmBfNR8H/fKfxx+u5rez6rqivU56Urz9ZfUXAAD//wMAUEsBAi0AFAAGAAgAAAAhAGLunWheAQAAkAQAABMAAAAAAAAA'
    'AAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYACAAAACEAtVUwI/QAAABMAgAACwAAAAAAAAAAAAAAAACXAwAA'
    'X3JlbHMvLnJlbHNQSwECLQAUAAYACAAAACEA8cXQoukCAAB5BgAADwAAAAAAAAAAAAAAAAC8BgAAeGwvd29ya2Jvb2sueG1sUEsB'
    'Ai0AFAAGAAgAAAAhAIE+lJfzAAAAugIAABoAAAAAAAAAAAAAAAAA0gkAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAi0A'
    'FAAGAAgAAAAhAAMTAIjuEQAAm4QAABgAAAAAAAAAAAAAAAAABQwAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbFBLAQItABQABgAI'
    'AAAAIQDBFxC+TgcAAMYgAAATAAAAAAAAAAAAAAAAACkeAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAi0AFAAGAAgAAAAhAGKZ4HNc'
    'CAAAXUEAAA0AAAAAAAAAAAAAAAAAqCUAAHhsL3N0eWxlcy54bWxQSwECLQAUAAYACAAAACEARTwoMCgCAAAFBgAAFAAAAAAAAAAA'
    'AAAAAAAvLgAAeGwvc2hhcmVkU3RyaW5ncy54bWxQSwECLQAUAAYACAAAACEAQpmzzUABAABtAgAAEQAAAAAAAAAAAAAAAACJMAAA'
    'ZG9jUHJvcHMvY29yZS54bWxQSwECLQAUAAYACAAAACEACTewYY4BAAARAwAAEAAAAAAAAAAAAAAAAAAAMwAAZG9jUHJvcHMvYXBw'
    'LnhtbFBLBQYAAAAACgAKAIACAADENQAAAAA='
)


class InvoiceConversionError(ValueError):
    """Raised when an uploaded workbook does not match the approved Invoice format."""


@dataclass(frozen=True)
class InvoiceFilenameMetadata:
    source_stem: str
    employee_query: str
    docu_name: str
    remarks: str


@dataclass(frozen=True)
class InvoiceSourceRecord:
    source_row: int
    docu_date: date
    inv_no: Any
    product_name: str
    prod_uom: Any
    inv_qty: Any


@dataclass(frozen=True)
class InvoiceConversionResult:
    output_bytes: bytes
    output_filename: str
    metadata: InvoiceFilenameMetadata
    process_date: date
    logp_no: str
    employee_name: str
    auditor_name: str
    records: tuple[InvoiceSourceRecord, ...]
    source_signature: str

    @property
    def row_count(self) -> int:
        return len(self.records)

    @property
    def first_invoice_date(self) -> date:
        return min(record.docu_date for record in self.records)

    @property
    def last_invoice_date(self) -> date:
        return max(record.docu_date for record in self.records)

    def preview_rows(self, limit: int = 200) -> list[dict[str, Any]]:
        shown = self.records[: max(0, int(limit))]
        return [
            {
                "trans_id": None,
                "login_date": self.process_date.isoformat(),
                "logp_no": self.logp_no,
                "employee_name": self.employee_name,
                "docu_date": record.docu_date.isoformat(),
                "docu_name": self.metadata.docu_name,
                "sold_no": None,
                "inv_no": record.inv_no,
                "pd_no": None,
                "prod_code": None,
                "prod_name": record.product_name,
                "prod_uom": record.prod_uom,
                "inv_qty": record.inv_qty,
                "disc_qty": None,
                "record_qty": None,
                "count_qty": None,
                "remarks": self.metadata.remarks,
                "auditor_name": self.auditor_name,
            }
            for record in shown
        ]


def _clean_invoice_uploaded_stem(filename: str) -> str:
    raw_name = Path(str(filename or "").strip()).name
    if not raw_name:
        raise InvoiceConversionError("The uploaded Invoice Excel filename is missing.")
    stem = Path(raw_name).stem
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem).strip()
    stem = re.sub(r"\s+", " ", stem)
    if not stem:
        raise InvoiceConversionError("The uploaded Invoice Excel filename is invalid.")
    return stem


def parse_invoice_filename(filename: str) -> InvoiceFilenameMetadata:
    stem = _clean_invoice_uploaded_stem(filename)
    docu_name_match = re.match(r"\s*(Invoice)\b", stem, flags=re.I)
    if not docu_name_match:
        raise InvoiceConversionError(
            "The filename must begin with Invoice, for example: "
            "Invoice Davido Good.xlsx."
        )

    normalized_stem = re.sub(r"[_-]+", " ", stem).casefold()
    normalized_stem = re.sub(r"\s+", " ", normalized_stem).strip()
    if re.search(r"\bsold\s*out\b", normalized_stem):
        remarks = "Sold Out"
    elif re.search(r"\bsotex\b", normalized_stem):
        remarks = "Sotex"
    elif re.search(r"\bgood\b", normalized_stem):
        remarks = "Good"
    elif re.search(r"\bregular\b", normalized_stem):
        remarks = "Good"
    else:
        raise InvoiceConversionError(
            "The filename must contain Good, Regular, Sotex, or Sold Out for remarks."
        )

    employee_part = stem[docu_name_match.end():]
    employee_part = re.sub(r"\bsold[\s_-]*out\b", " ", employee_part, flags=re.I)
    employee_part = re.sub(r"\b(?:good|regular|sotex)\b", " ", employee_part, flags=re.I)
    employee_part = re.sub(
        r"(?<!\d)(?:0?[1-9]|1[0-2])[-_/](?:0?[1-9]|[12]\d|3[01])[-_/]\d{4}(?!\d)",
        " ",
        employee_part,
    )
    employee_part = re.sub(r"[_-]+", " ", employee_part)
    employee_query = " ".join(employee_part.split()).strip(" ,")
    if not employee_query:
        raise InvoiceConversionError(
            "The employee name was not found in the Invoice filename."
        )

    return InvoiceFilenameMetadata(
        source_stem=stem,
        employee_query=employee_query,
        docu_name="Invoice",
        remarks=remarks,
    )


def _resolve_invoice_employee_name(
    source_name: str,
    employee_records: Iterable[dict[str, Any]],
) -> str:
    source_text = " ".join(str(source_name or "").split()).strip()
    if not source_text:
        raise InvoiceConversionError("The employee name in the filename is blank.")

    records = list(employee_records or [])
    if not records:
        raise InvoiceConversionError(
            "Master Data Employees is required to resolve the Invoice employee name."
        )

    source_tokens_list = _person_tokens(source_text)
    source_tokens = set(source_tokens_list)
    scored: list[tuple[int, str]] = []
    for record in records:
        official_name = " ".join(str(record.get("name") or "").split()).strip()
        if not official_name:
            continue
        variants = [official_name, *(record.get("aliases") or [])]
        best_score = -1
        for variant in variants:
            variant_tokens_list = _person_tokens(variant)
            variant_tokens = set(variant_tokens_list)
            if not variant_tokens:
                continue
            source_norm = " ".join(source_tokens_list)
            variant_norm = " ".join(variant_tokens_list)
            if source_norm == variant_norm:
                best_score = max(best_score, 1000)
                continue
            if source_tokens and source_tokens.issubset(variant_tokens):
                score = 500 + len(source_tokens) * 10
                if source_tokens_list[-1:] == variant_tokens_list[-1:]:
                    score += 25
                best_score = max(best_score, score)
        if best_score >= 0:
            scored.append((best_score, official_name))

    if not scored:
        raise InvoiceConversionError(
            f'Employee "{source_text}" from the filename could not be matched to Master Data Employees.'
        )

    top_score = max(score for score, _ in scored)
    top_names = sorted(
        {name for score, name in scored if score == top_score}, key=str.casefold
    )
    if len(top_names) != 1:
        raise InvoiceConversionError(
            f'Employee "{source_text}" matched multiple Master Data employees: '
            + ", ".join(top_names)
        )
    return top_names[0]


def _coerce_invoice_date(value: Any, workbook: Any, source_row: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(value, workbook.epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception as exc:
            raise InvoiceConversionError(
                f"Invoice date in column C is invalid at source row {source_row}."
            ) from exc
    text = " ".join(str(value or "").split()).strip()
    for pattern in ("%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise InvoiceConversionError(
        f"Invoice date in column C is blank or invalid at source row {source_row}."
    )


def _normalize_logp_number(value: Any) -> str:
    if isinstance(value, bool) or value in (None, ""):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split()).strip()




def _normalize_invoice_number(value: Any) -> Any:
    if isinstance(value, bool) or value in (None, ""):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = " ".join(str(value).split()).strip()
    if re.fullmatch(r"[+-]?\d+", text):
        return int(text)
    return text


def _divide_invoice_quantity_by_ten(value: Any, source_row: int) -> Any:
    """Convert a blister quantity to boxes while retaining its sign."""
    if isinstance(value, bool) or value in (None, ""):
        raise InvoiceConversionError(
            f"Invoice quantity in column F is blank or invalid at source row {source_row}; "
            "a blister UOM requires a numeric quantity."
        )

    if isinstance(value, (int, float)):
        numeric_value = float(value)
    else:
        text = "".join(str(value).split()).replace(",", "")
        try:
            numeric_value = float(text)
        except (TypeError, ValueError) as exc:
            raise InvoiceConversionError(
                f"Invoice quantity in column F is not numeric at source row {source_row}; "
                "a blister UOM requires a numeric quantity."
            ) from exc

    converted = numeric_value / 10
    if converted.is_integer():
        return int(converted)
    return converted


def _normalize_invoice_uom_and_quantity(
    uom_value: Any, quantity_value: Any, source_row: int
) -> tuple[Any, Any]:
    """Convert blister/blisters to box and divide quantity by 10."""
    if uom_value in (None, ""):
        return None, quantity_value

    normalized_uom = " ".join(str(uom_value).split()).strip()
    if normalized_uom.casefold() in {"blister", "blisters"}:
        return "box", _divide_invoice_quantity_by_ten(quantity_value, source_row)
    return uom_value, quantity_value


def extract_invoice_records(
    excel_bytes: bytes,
) -> tuple[str, tuple[InvoiceSourceRecord, ...]]:
    if not excel_bytes:
        raise InvoiceConversionError("The uploaded Invoice Excel file is empty.")

    try:
        workbook = load_workbook(BytesIO(excel_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise InvoiceConversionError(
            "The uploaded file could not be opened as a valid .xlsx workbook."
        ) from exc

    worksheet = workbook.active
    logp_numbers: list[str] = []
    records: list[InvoiceSourceRecord] = []

    for row_index in range(1, worksheet.max_row + 1):
        marker = " ".join(str(worksheet.cell(row_index, 1).value or "").split()).upper()
        if marker == "LOGP":
            logp_number = _normalize_logp_number(worksheet.cell(row_index, 2).value)
            if logp_number:
                logp_numbers.append(logp_number)
            continue
        if marker != "INV":
            continue

        product_raw = str(worksheet.cell(row_index, 5).value or "").strip()
        if not product_raw:
            raise InvoiceConversionError(
                f"Product name in column E is blank at source row {row_index}."
            )
        inv_no = _normalize_invoice_number(worksheet.cell(row_index, 10).value)
        if inv_no in (None, ""):
            raise InvoiceConversionError(
                f"Invoice number in column J is blank at source row {row_index}."
            )
        prod_uom, inv_qty = _normalize_invoice_uom_and_quantity(
            worksheet.cell(row_index, 7).value,
            worksheet.cell(row_index, 6).value,
            row_index,
        )
        records.append(
            InvoiceSourceRecord(
                source_row=row_index,
                docu_date=_coerce_invoice_date(
                    worksheet.cell(row_index, 3).value, workbook, row_index
                ),
                inv_no=inv_no,
                product_name=_remove_apostrophes(product_raw),
                prod_uom=prod_uom,
                inv_qty=inv_qty,
            )
        )

    unique_logp_numbers = list(dict.fromkeys(logp_numbers))
    if not unique_logp_numbers:
        raise InvoiceConversionError(
            "No LOGP number was found in column B beside a LOGP marker in column A."
        )
    if len(unique_logp_numbers) != 1:
        raise InvoiceConversionError(
            "Multiple LOGP numbers were found in the source file: "
            + ", ".join(unique_logp_numbers)
        )
    if not records:
        raise InvoiceConversionError(
            "No Invoice rows were captured. Column A must contain INV for convertible rows."
        )
    return unique_logp_numbers[0], tuple(records)


def _load_invoice_template(template_path: Path | None = None):
    path = Path(template_path or INVOICE_TEMPLATE_PATH)
    try:
        if path.exists():
            workbook = load_workbook(path)
        else:
            template_bytes = base64.b64decode(INVOICE_TEMPLATE_XLSX_BASE64)
            workbook = load_workbook(BytesIO(template_bytes))
    except Exception as exc:
        raise InvoiceConversionError(
            "The Invoice conversion template could not be opened."
        ) from exc

    worksheet = workbook.active
    headers = [worksheet.cell(1, column).value for column in range(1, 19)]
    if headers != INVOICE_OUTPUT_HEADERS:
        raise InvoiceConversionError(
            "The Invoice conversion template headers do not match the approved format."
        )
    return workbook, worksheet


def _write_invoice_output_rows(
    worksheet: Any,
    records: Iterable[InvoiceSourceRecord],
    metadata: InvoiceFilenameMetadata,
    process_date: date,
    logp_no: str,
    employee_name: str,
    auditor_name: str,
) -> int:
    records = tuple(records)
    style_prototypes = [copy(worksheet.cell(2, column)._style) for column in range(1, 19)]
    alignment_prototypes = [
        copy(worksheet.cell(2, column).alignment) for column in range(1, 19)
    ]
    protection_prototypes = [
        copy(worksheet.cell(2, column).protection) for column in range(1, 19)
    ]

    clear_through = max(worksheet.max_row, len(records) + 1)
    for row_index in range(2, clear_through + 1):
        for column_index in range(1, 19):
            worksheet.cell(row_index, column_index).value = None

    login_date = datetime.combine(process_date, datetime.min.time())
    logp_value: Any = int(logp_no) if logp_no.isdigit() else logp_no

    for row_index, record in enumerate(records, start=2):
        values = [
            None,
            login_date,
            logp_value,
            employee_name,
            datetime.combine(record.docu_date, datetime.min.time()),
            metadata.docu_name,
            None,
            record.inv_no,
            None,
            None,
            record.product_name,
            record.prod_uom,
            record.inv_qty,
            None,
            None,
            None,
            metadata.remarks,
            auditor_name,
        ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column_index)
            cell._style = copy(style_prototypes[column_index - 1])
            cell.alignment = copy(alignment_prototypes[column_index - 1])
            cell.protection = copy(protection_prototypes[column_index - 1])
            cell.value = value
            cell.number_format = r"yyyy\-mm\-dd" if column_index in (2, 5) else "General"

    target_last_row = len(records) + 1
    if worksheet.max_row > target_last_row:
        worksheet.delete_rows(target_last_row + 1, worksheet.max_row - target_last_row)
    return len(records)


def _safe_invoice_output_filename(
    metadata: InvoiceFilenameMetadata, employee_name: str
) -> str:
    surname = _person_tokens(employee_name)[-1] if _person_tokens(employee_name) else "EMPLOYEE"
    raw = f"Invoice for upload {surname.title()} {metadata.remarks}.xlsx"
    return re.sub(r'[<>:"/\\|?*]+', "_", raw)


def build_invoice_conversion(
    excel_bytes: bytes,
    filename: str,
    auditor_name: str,
    employee_records: Iterable[dict[str, Any]],
    *,
    process_date: date | None = None,
    template_path: Path | None = None,
) -> InvoiceConversionResult:
    conversion_date = process_date or philippine_today()
    clean_auditor_name = " ".join(str(auditor_name or "").split()).strip()
    if not clean_auditor_name:
        raise InvoiceConversionError(
            "The signed-in user's full name is required for auditor_name."
        )

    metadata = parse_invoice_filename(filename)
    employee_name = _resolve_invoice_employee_name(
        metadata.employee_query, employee_records
    )
    logp_no, records = extract_invoice_records(excel_bytes)
    workbook, worksheet = _load_invoice_template(template_path)
    _write_invoice_output_rows(
        worksheet,
        records,
        metadata,
        conversion_date,
        logp_no,
        employee_name,
        clean_auditor_name,
    )

    if workbook.calculation is not None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    buffer = BytesIO()
    workbook.save(buffer)
    output_bytes = buffer.getvalue()

    try:
        verification_book = load_workbook(
            BytesIO(output_bytes), data_only=False, read_only=True
        )
        verification_sheet = verification_book.active
        if verification_sheet.max_row != len(records) + 1:
            raise InvoiceConversionError(
                "Converted row count did not match the captured Invoice row count."
            )
        verified_headers = [
            verification_sheet.cell(1, column).value for column in range(1, 19)
        ]
        if verified_headers != INVOICE_OUTPUT_HEADERS:
            raise InvoiceConversionError(
                "The converted Invoice headers failed the final integrity check."
            )
        for output_row, source_record in enumerate(records, start=2):
            converted_date = verification_sheet.cell(output_row, 5).value
            if isinstance(converted_date, datetime):
                converted_date = converted_date.date()
            if converted_date != source_record.docu_date:
                raise InvoiceConversionError(
                    "An Invoice document date did not match its own source row in column C."
                )
    except InvoiceConversionError:
        raise
    except Exception as exc:
        raise InvoiceConversionError(
            "The converted Invoice Excel file failed the final workbook integrity check."
        ) from exc

    return InvoiceConversionResult(
        output_bytes=output_bytes,
        output_filename=_safe_invoice_output_filename(metadata, employee_name),
        metadata=metadata,
        process_date=conversion_date,
        logp_no=logp_no,
        employee_name=employee_name,
        auditor_name=clean_auditor_name,
        records=records,
        source_signature=hashlib.sha256(excel_bytes).hexdigest(),
    )


def render_invoice_conversion_page(
    user: dict[str, Any],
    employee_records: Iterable[dict[str, Any]],
) -> None:
    import pandas as pd
    import streamlit as st

    st.markdown(
        """
        <style>
        .iars-invoice-hero {
            border: 1px solid #DDE5EF;
            border-radius: 16px;
            padding: 1.05rem 1.15rem;
            margin: 0 0 .9rem 0;
            background: linear-gradient(135deg, #F8FAFD 0%, #FFFFFF 58%, #FFF9EB 100%);
            box-shadow: 0 8px 24px rgba(6,26,54,.06);
        }
        .iars-invoice-hero h2 { margin: 0; color: #061A36; font-size: 1.35rem; }
        .iars-invoice-hero p { margin: .28rem 0 0; color: #667085; font-size: .88rem; }
        .iars-invoice-route {
            display: grid;
            grid-template-columns: 1fr auto 1fr auto 1fr;
            align-items: center;
            gap: .55rem;
            margin-top: .85rem;
        }
        .iars-invoice-route div {
            min-height: 64px;
            border: 1px solid #E4EAF2;
            border-radius: 12px;
            padding: .65rem .72rem;
            background: rgba(255,255,255,.9);
        }
        .iars-invoice-route strong { display:block; color:#061A36; font-size:.86rem; }
        .iars-invoice-route span { color:#667085; font-size:.74rem; line-height:1.25; }
        .iars-invoice-route b { color:#C78B12; font-size:1.1rem; }
        @media (max-width: 760px) {
            .iars-invoice-route { grid-template-columns: 1fr; }
            .iars-invoice-route b { display:none; }
        }
        </style>
        <div class="iars-invoice-hero">
          <h2>Sales Personnel Invoice Conversion</h2>
          <p>Convert SAP Invoice rows into the approved Sales Personnel upload template.</p>
          <div class="iars-invoice-route">
            <div><strong>1. SAP Invoice Excel</strong><span>Upload the original Invoice file with remarks in its filename.</span></div>
            <b>→</b>
            <div><strong>2. IARS Mapping</strong><span>Capture INV rows, retain each Invoice date and convert blister quantities to boxes.</span></div>
            <b>→</b>
            <div><strong>3. Compatible Output</strong><span>Download the approved 18-column Invoice template.</span></div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        st.markdown("### Upload SAP Invoice File")
        st.caption(
            "Filename examples: Invoice Davido Good.xlsx · Invoice Davido Sotex.xlsx · "
            "Invoice Davido Sold Out.xlsx"
        )
        uploaded_file = st.file_uploader(
            "SAP Sales Personnel Invoice Excel",
            type=["xlsx"],
            key="sales_invoice_excel_uploader_v4_5_15",
            help=(
                "Only rows marked INV in column A are converted. The source file remains unchanged."
            ),
        )

    if uploaded_file is None:
        st.info(
            "Upload an Invoice .xlsx file. IARS will capture each INV row's date from column C, "
            "product from E, quantity from F, UOM from G and invoice number from J. Blister UOM is converted to box and quantity is divided by 10."
        )
        return

    auditor_name = str(user.get("full_name") or user.get("username") or "").strip()
    process_date = philippine_today()

    try:
        with st.spinner("Validating and converting the Sales Personnel Invoice file…"):
            result = build_invoice_conversion(
                uploaded_file.getvalue(),
                uploaded_file.name,
                auditor_name,
                employee_records,
                process_date=process_date,
            )
    except InvoiceConversionError as exc:
        st.error(str(exc))
        return
    except Exception as exc:
        st.error(f"Invoice conversion failed: {exc}")
        return

    metadata = result.metadata
    metric_columns = st.columns(4)
    metric_columns[0].metric("Invoice Rows", f"{result.row_count:,}")
    metric_columns[1].metric("LOGP No.", result.logp_no)
    metric_columns[2].metric(
        "Invoice Dates",
        (
            result.first_invoice_date.isoformat()
            if result.first_invoice_date == result.last_invoice_date
            else f"{result.first_invoice_date.isoformat()} to {result.last_invoice_date.isoformat()}"
        ),
    )
    metric_columns[3].metric("Remarks", metadata.remarks)

    st.success(
        f"Conversion completed for {result.row_count:,} INV rows. "
        "Each row retained its own Invoice date from column C, negative quantities were retained, "
        "product apostrophes were removed, blister UOM quantities were converted to boxes, and the output passed the workbook integrity check."
    )

    with st.expander("Conversion Details", expanded=True):
        details = pd.DataFrame(
            [
                ["Login Date", result.process_date.isoformat()],
                ["Source Filename", uploaded_file.name],
                ["LOGP No.", result.logp_no],
                ["Document Name", metadata.docu_name],
                ["Invoice Date Range", f"{result.first_invoice_date.isoformat()} to {result.last_invoice_date.isoformat()}"],
                ["Remarks", metadata.remarks],
                ["Employee Name", result.employee_name],
                ["Auditor Name", result.auditor_name],
            ],
            columns=["Field", "Generated Value"],
        )
        st.dataframe(details, hide_index=True, width="stretch")

    st.markdown("### Converted Data Preview")
    preview = pd.DataFrame(result.preview_rows(limit=200))
    st.dataframe(preview, hide_index=True, width="stretch", height=390)
    if result.row_count > len(preview):
        st.caption(
            f"Showing the first {len(preview):,} of {result.row_count:,} converted rows."
        )

    st.download_button(
        "⬇️ Download Converted Invoice Excel",
        data=result.output_bytes,
        file_name=result.output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"invoice_download_{result.source_signature[:16]}",
        type="primary",
        width="stretch",
    )
    st.caption(
        "Output format: approved 18 columns · dates in yyyy-mm-dd · each docu_date comes from "
        "its own INV source row in column C · quantity/UOM come from columns F/G · "
        "blister or blisters is converted to box and quantity is divided by 10 · "
        "trans_id, sold_no, pd_no, prod_code, disc_qty, record_qty and count_qty remain blank."
    )
